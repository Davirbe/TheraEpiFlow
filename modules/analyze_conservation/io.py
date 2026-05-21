"""I/O for analyze_conservation: FASTA loading and the conservation-summary +
mutations XLSX writers (with the conservation colour palette)."""

from pathlib import Path

import pandas as pd
from Bio import SeqIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .core import _LENGTH_TOLERANCE

_FILL_GRAY      = PatternFill("solid", fgColor="D9D9D9")
_FILL_PERFECT   = PatternFill("solid", fgColor="00B050")
_FILL_HIGH      = PatternFill("solid", fgColor="92D050")
_FILL_MODERATE  = PatternFill("solid", fgColor="FFFF99")
_FILL_LOW       = PatternFill("solid", fgColor="FF9999")
_FILL_UNKNOWN   = PatternFill("solid", fgColor="D9D9D9")

_LABEL_FILL = {
    "perfect":              _FILL_PERFECT,
    "high":                 _FILL_HIGH,
    "moderate":             _FILL_MODERATE,
    "low":                  _FILL_LOW,
    "conservation_unknown": _FILL_UNKNOWN,
}
_VERDICT_FILL = {
    "likely_lost":     PatternFill("solid", fgColor="FF9999"),
    "excellent_match": PatternFill("solid", fgColor="92D050"),
    "tolerated":       PatternFill("solid", fgColor="FFFF99"),
}

_VERDICT_ORDER = {"excellent_match": 0, "tolerated": 1, "likely_lost": 2}
# ── FASTA loading ─────────────────────────────────────────────────────────────

def load_fasta_sequences(
    fasta_path: Path, ref_length: int = 0, apply_length_filter: bool = True,
) -> tuple[list, int]:
    """Returns (records, n_excluded). When ref_length > 0 and apply_length_filter is True,
    filters to ±_LENGTH_TOLERANCE of the reference. Set apply_length_filter=False to keep
    every variant regardless of length (e.g. to include partial/fragment sequences)."""
    records = list(SeqIO.parse(str(fasta_path), "fasta"))
    if ref_length <= 0 or not apply_length_filter:
        return records, 0
    lo = max(1, int(ref_length * (1 - _LENGTH_TOLERANCE)))
    hi = int(ref_length * (1 + _LENGTH_TOLERANCE))
    kept = [r for r in records if lo <= len(r.seq) <= hi]
    return kept, len(records) - len(kept)


# ── XLSX writers ──────────────────────────────────────────────────────────────

def write_conservation_summary_xlsx(df: pd.DataFrame, output_path: Path):
    """
    IEDB-style XLSX: header gray, no row backgrounds, only `conservation_label`
    cell coloured. Single sheet, frozen header.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Conservation Summary"

    columns = list(df.columns)

    for col_idx, col_name in enumerate(columns, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = _FILL_GRAY

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            if col_name == "conservation_label":
                cell.fill      = _LABEL_FILL.get(row[col_name], _FILL_UNKNOWN)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "#":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_width  = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].tolist()),
        )
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    wb.save(str(output_path))


_MUTATIONS_COLUMNS = [
    "peptide", "length", "variant_accession", "variant_window",
    "identity", "n_mutations", "mutations", "mutation_positions",
    "anchor_hit", "blosum62_min", "mhc_verdict", "alleles_united",
]


def write_mutations_xlsx(mutation_records: list[dict], output_path: Path):
    """
    Per (epitope, variant) breakdown for variants with 1 or 2 mutations.
    Rows coloured by `mhc_verdict`; sorted excellent → tolerated → likely_lost.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutation Detail"

    for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = _FILL_GRAY

    if not mutation_records:
        ws.cell(row=2, column=1, value="(no variants with 1 or 2 mutations)")
        ws.freeze_panes = "A2"
        wb.save(str(output_path))
        return

    sorted_records = sorted(
        mutation_records,
        key=lambda r: (
            _VERDICT_ORDER.get(r["mhc_verdict"], 99),
            r["peptide"],
            r["n_mutations"],
        ),
    )

    left_aligned = {"alleles_united", "mutations", "peptide", "variant_window", "variant_accession"}
    for row_idx, record in enumerate(sorted_records, start=2):
        row_fill = _VERDICT_FILL.get(record["mhc_verdict"], _FILL_UNKNOWN)
        for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name))
            cell.fill      = row_fill
            cell.alignment = Alignment(
                horizontal="left" if col_name in left_aligned else "center",
                vertical="center",
            )

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
        col_letter      = ws.cell(row=1, column=col_idx).column_letter
        column_values   = [str(r.get(col_name, "")) for r in sorted_records]
        max_width       = max([len(col_name)] + [len(v) for v in column_values])
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    wb.save(str(output_path))


