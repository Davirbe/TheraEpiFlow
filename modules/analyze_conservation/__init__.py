"""
analyze_conservation step.

Takes the ★ representatives from select_representatives and measures how
faithfully each epitope appears in the variant sequences produced by
search_variants. Conservation is quantified via sliding window identity:
for each variant sequence the best-matching window of len(peptide) residues
is found, and its per-position match ratio is recorded.

The analysis threshold (default 1.0 = exact match) is configurable per
project. It determines which variants are labelled "passed" vs "failed".
The conservation_label and all row/cell colours are based on
mean_max_identity and never change regardless of the chosen threshold.

The step is qualitative — no epitopes are removed.

Inputs (clusters/):
    CLUSTER_REPR_{track_id}.csv          select_representatives output

Inputs (variants/):
    VARIANTS_{track_id}.fasta            search_variants output or user-supplied

Outputs (conservation/):
    CONSERVATION_{track_id}.csv             IEDB-style summary, one row per ★ rep
    CONSERVATION_{track_id}.xlsx            same table, IEDB-style minimal styling
    CONSERVATION_HEATMAP_{track_id}.png     dual-panel: position cons + identity tiers
    CONSERVATION_MUTATIONS_{track_id}.xlsx  per (epitope, variant) pair with 1-2 muts
    CONSERVATION_AUDIT_{track_id}.json      run metadata + verdict counts

External dependencies:
    Biopython >= 1.79 (for Bio.SeqIO and substitution_matrices.load("BLOSUM62")).
    The BLOSUM62 matrix used for the mutation verdict is the canonical
    Henikoff & Henikoff 1992 matrix as packaged in Biopython — no other source.

Citations:
    Henikoff S, Henikoff JG. PNAS. 1992;89(22):10915-10919. (BLOSUM62)
"""

import datetime
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Optional

import pandas as pd
from Bio import SeqIO
from Bio.Align import substitution_matrices
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rich import box
from rich.panel import Panel
from rich.table import Table
from rich.progress import (
    Progress, SpinnerColumn, TextColumn,
    BarColumn, MofNCompleteColumn, TimeElapsedColumn,
)

import config
from modules.base_step import BaseTrackStep
from utils.console import console, is_interactive_session
from utils.naming import (
    COLUMN_ALLELES_UNITED,
    COLUMN_BEST_REPRESENTATIVE,
    COLUMN_NUM_ALLELES_UNITED,
    COLUMN_PEPTIDE,
    STAR_MARKER,
    get_step_filename,
)
from utils.project_manager import save_project_config

# ── Colour palette ────────────────────────────────────────────────────────────

_FILL_GRAY      = PatternFill("solid", fgColor="D9D9D9")
_FILL_PERFECT   = PatternFill("solid", fgColor="00B050")
_FILL_HIGH      = PatternFill("solid", fgColor="92D050")
_FILL_MODERATE  = PatternFill("solid", fgColor="FFFF99")
_FILL_LOW       = PatternFill("solid", fgColor="FF9999")
_FILL_UNKNOWN   = PatternFill("solid", fgColor="D9D9D9")

_LABEL_FILL = {
    "perfect":              _FILL_PERFECT,
    "high":                 _FILL_HIGH,
    "moderate":             _FILL_MODERATE,
    "low":                  _FILL_LOW,
    "conservation_unknown": _FILL_UNKNOWN,
}

_LABEL_HEX = {
    "perfect":              "#00B050",
    "high":                 "#92D050",
    "moderate":             "#FFFF99",
    "low":                  "#FF9999",
    "conservation_unknown": "#D9D9D9",
}

_VERDICT_FILL = {
    "likely_lost":     PatternFill("solid", fgColor="FF9999"),
    "excellent_match": PatternFill("solid", fgColor="92D050"),
    "tolerated":       PatternFill("solid", fgColor="FFFF99"),
}

_VERDICT_ORDER = {"excellent_match": 0, "tolerated": 1, "likely_lost": 2}

_ANCHOR_BORDER_HEX = "#2E75B6"

# Accept variants whose length is within ±_LENGTH_TOLERANCE of the reference.
_LENGTH_TOLERANCE = 0.25

# Max mutations to qualify a (peptide, variant) pair for the MUTATIONS XLSX
# and to qualify an epitope for inclusion in the heatmap PNG.
_MAX_MUTATIONS_FOR_HEATMAP = 2

# BLOSUM62 substitution matrix from Biopython. Score >= 0 = conservative.
_BLOSUM62_MATRIX = substitution_matrices.load("BLOSUM62")


# ── Substitution scoring + anchor helpers ─────────────────────────────────────

def _blosum62_score(ref_aa: str, var_aa: str) -> int:
    """BLOSUM62 substitution score. >=0 = conservative substitution."""
    try:
        return int(_BLOSUM62_MATRIX[(ref_aa, var_aa)])
    except (KeyError, IndexError):
        return -4


def _anchor_positions(peptide_length: int) -> tuple[int, int]:
    """Returns (p2_idx, pomega_idx) 0-based for MHC-I anchor positions."""
    return 1, peptide_length - 1


def _classify_mhc_verdict(n_mutations: int, anchor_hit: bool, blosum62_min: int) -> str:
    """
    Heuristic verdict for variant tolerance based on anchor + BLOSUM62.

    likely_lost     → any mutation in P2 or PΩ
    excellent_match → 1 mutation, non-anchor, BLOSUM62 ≥ 0
    tolerated       → 1-2 non-anchor mutations otherwise
    """
    if anchor_hit:
        return "likely_lost"
    if n_mutations == 1 and blosum62_min >= 0:
        return "excellent_match"
    return "tolerated"


def _format_pct_fraction(numerator: int, denominator: int) -> str:
    """Returns "XX.XX% (n/N)". Empty denominator → "—"."""
    if denominator <= 0:
        return "—"
    pct = numerator / denominator * 100
    return f"{pct:.2f}% ({numerator}/{denominator})"


# ── Core computation ──────────────────────────────────────────────────────────

def compute_max_window_identity(peptide: str, sequence: str) -> tuple[float, str]:
    """
    Slides a window of len(peptide) over sequence, returning the maximum
    per-position identity (0.0–1.0) and the corresponding window string.
    Returns (0.0, "") when the sequence is shorter than the peptide.
    """
    peptide_length = len(peptide)
    best_identity  = 0.0
    best_window    = ""
    for start in range(len(sequence) - peptide_length + 1):
        window   = sequence[start : start + peptide_length]
        identity = sum(ref == var for ref, var in zip(peptide, window)) / peptide_length
        if identity > best_identity:
            best_identity = identity
            best_window   = window
    return best_identity, best_window


def classify_conservation_label(mean_max_identity: float, n_variants_total: int) -> str:
    if n_variants_total == 0:
        return "conservation_unknown"
    if mean_max_identity == 1.0:
        return "perfect"
    if mean_max_identity >= config.CONSERVATION_HIGH_THRESHOLD:
        return "high"
    if mean_max_identity >= config.CONSERVATION_MODERATE_THRESHOLD:
        return "moderate"
    return "low"


def _variant_label(record) -> str:
    """Returns a short 'accession(organism)' label from a SeqRecord."""
    desc_parts = record.description.strip().split(" ", 1)
    organism_short = (
        desc_parts[1].split("|")[0].strip()[:30]
        if len(desc_parts) > 1
        else record.id
    )
    return f"{record.id}({organism_short})"


def compute_epitope_conservation(
    peptide: str,
    records: list,
    analysis_threshold: float,
) -> tuple[dict, list[tuple[str, float, str]]]:
    """
    Computes all conservation metrics for a single peptide against every
    variant record. Returns (metrics_dict, alignment_tuples).

    alignment_tuples: [(variant_label, identity, best_window), ...] for every
    record, re-used downstream for mutation records and the position heatmap.

    Metrics carry RAW COUNTS (not pre-formatted percentages). The CSV/XLSX
    writers compose "XX.XX% (n/N)" strings from these.
    """
    n_total: int = len(records)
    alignment_tuples: list[tuple[str, float, str]] = []

    if n_total == 0:
        empty_metrics = {
            "n_variants_used":      0,
            "n_passed_threshold":   0,
            "n_exact_match":        0,
            "n_identity_90":        0,
            "n_identity_80":        0,
            "analysis_threshold":   analysis_threshold,
            "mean_max_identity":    0.0,
            "min_max_identity":     0.0,
            "max_max_identity":     0.0,
            "conservation_label":   "conservation_unknown",
        }
        return empty_metrics, alignment_tuples

    per_variant_identities: list[float] = []
    for record in records:
        sequence = str(record.seq).upper()
        identity, window = compute_max_window_identity(peptide, sequence)
        per_variant_identities.append(identity)
        alignment_tuples.append((_variant_label(record), identity, window))

    mean_id  = sum(per_variant_identities) / n_total
    min_id   = min(per_variant_identities)
    max_id   = max(per_variant_identities)
    n_exact  = sum(1 for i in per_variant_identities if i == 1.0)
    n_90pct  = sum(1 for i in per_variant_identities if i >= 0.90)
    n_80pct  = sum(1 for i in per_variant_identities if i >= 0.80)
    n_passed = sum(1 for i in per_variant_identities if i >= analysis_threshold)

    metrics = {
        "n_variants_used":      n_total,
        "n_passed_threshold":   n_passed,
        "n_exact_match":        n_exact,
        "n_identity_90":        n_90pct,
        "n_identity_80":        n_80pct,
        "analysis_threshold":   analysis_threshold,
        "mean_max_identity":    round(mean_id, 4),
        "min_max_identity":     round(min_id, 4),
        "max_max_identity":     round(max_id, 4),
        "conservation_label":   classify_conservation_label(mean_id, n_total),
    }
    return metrics, alignment_tuples


def compute_mutation_records(
    peptide: str,
    alignment_tuples: list[tuple[str, float, str]],
    alleles_united: str,
) -> list[dict]:
    """
    For each (variant, peptide) pair where the best window has 1 or 2
    substitutions, builds a record capturing position(s), BLOSUM62 score, and
    the MHC verdict. Pairs with 0 mutations or >2 mutations are skipped.
    """
    if not peptide or not alignment_tuples:
        return []

    peptide_length      = len(peptide)
    p2_idx, pomega_idx  = _anchor_positions(peptide_length)
    anchor_position_set = {p2_idx, pomega_idx}

    records = []
    for variant_label, _identity, window in alignment_tuples:
        if len(window) != peptide_length:
            continue

        mutation_triples = [
            (pos, ref_aa, var_aa)
            for pos, (ref_aa, var_aa) in enumerate(zip(peptide, window))
            if ref_aa != var_aa
        ]
        n_mutations = len(mutation_triples)
        if n_mutations < 1 or n_mutations > _MAX_MUTATIONS_FOR_HEATMAP:
            continue

        anchor_hit    = any(pos in anchor_position_set for pos, _, _ in mutation_triples)
        blosum_scores = [_blosum62_score(ref, var) for _, ref, var in mutation_triples]
        blosum_min    = min(blosum_scores)
        verdict       = _classify_mhc_verdict(n_mutations, anchor_hit, blosum_min)

        accession      = variant_label.split("(")[0]
        mutations_str  = "; ".join(f"P{pos + 1}:{ref}→{var}" for pos, ref, var in mutation_triples)
        positions_str  = ",".join(str(pos + 1) for pos, _, _ in mutation_triples)
        n_match        = peptide_length - n_mutations

        records.append({
            "peptide":            peptide,
            "length":             peptide_length,
            "variant_accession":  accession,
            "variant_window":     window,
            "identity":           _format_pct_fraction(n_match, peptide_length),
            "n_mutations":        n_mutations,
            "mutations":          mutations_str,
            "mutation_positions": positions_str,
            "anchor_hit":         "Y" if anchor_hit else "N",
            "blosum62_min":       blosum_min,
            "mhc_verdict":        verdict,
            "alleles_united":     alleles_united,
        })

    return records


def _has_variant_within_max_mut(peptide: str, alignment_tuples: list) -> bool:
    """True when at least one variant has ≤ _MAX_MUTATIONS_FOR_HEATMAP differences."""
    peptide_length = len(peptide)
    for _, _, window in alignment_tuples:
        if len(window) != peptide_length:
            continue
        n_mut = sum(a != b for a, b in zip(peptide, window))
        if n_mut <= _MAX_MUTATIONS_FOR_HEATMAP:
            return True
    return False


def _align_position_stats_to_anchor(
    pos_stats: list[dict],
    peptide_length: int,
    max_length: int,
) -> tuple[list, list]:
    """
    Maps a peptide's per-position stats into a max_length-wide grid by
    anchor-aligning both ends: N-terminal half left-justified, C-terminal
    half right-justified. Middle positions (the MHC-I "bulge") become None.

    Guarantees: P2 of every peptide lands at column 1, PΩ at column
    max_length - 1 — so anchor borders line up across mixed lengths.
    """
    aligned_pcts:  list = [None] * max_length
    aligned_annot: list = [""]   * max_length
    n_term_half = peptide_length // 2
    for pep_idx in range(peptide_length):
        col = (
            pep_idx
            if pep_idx < n_term_half
            else max_length - (peptide_length - pep_idx)
        )
        pct = pos_stats[pep_idx]["conservation_pct"]
        aligned_pcts[col] = pct
        if pct < 100.0:
            aligned_annot[col] = (
                f"{int(round(pct))}%\n{pos_stats[pep_idx]['top_mutation']}"
            )
    return aligned_pcts, aligned_annot


def _build_anchor_aligned_xtick_labels(max_length: int) -> list[str]:
    """
    X-axis labels matching the layout produced by
    _align_position_stats_to_anchor: N-terminal half uses absolute
    numbering (P1, P2, ...); C-terminal half uses offset-from-end
    (..., PΩ-1, PΩ).
    """
    n_term_half = max_length // 2
    n_term_labels = [f"P{i + 1}" for i in range(n_term_half)]
    c_term_count  = max_length - n_term_half
    c_term_labels = [
        ("PΩ" if (c_term_count - 1 - k) == 0 else f"PΩ-{c_term_count - 1 - k}")
        for k in range(c_term_count)
    ]
    return n_term_labels + c_term_labels


def _compute_position_stats(peptide: str, alignment_tuples: list[tuple]) -> list[dict]:
    """
    For each position in the peptide returns:
      conservation_pct  — % of variants with the same AA as reference (0–100)
      top_mutation      — most common substitution string e.g. "Q→R", or "" if 100%
    """
    n_variants = len(alignment_tuples)
    stats = []
    for pos_idx, ref_aa in enumerate(peptide):
        if n_variants == 0:
            stats.append({"conservation_pct": 0.0, "top_mutation": ""})
            continue
        n_match: int = 0
        mutation_counter: Counter = Counter()
        for _, _, window in alignment_tuples:
            if len(window) > pos_idx:
                var_aa = window[pos_idx]
                if var_aa == ref_aa:
                    n_match += 1
                else:
                    mutation_counter[f"{ref_aa}→{var_aa}"] += 1
        pct     = n_match / n_variants * 100
        top_mut = mutation_counter.most_common(1)[0][0] if mutation_counter else ""
        stats.append({"conservation_pct": pct, "top_mutation": top_mut})
    return stats


# ── FASTA loading ─────────────────────────────────────────────────────────────

def load_fasta_sequences(fasta_path: Path, ref_length: int = 0) -> tuple[list, int]:
    """Returns (records, n_excluded). When ref_length > 0, filters to ±_LENGTH_TOLERANCE."""
    records = list(SeqIO.parse(str(fasta_path), "fasta"))
    if ref_length <= 0:
        return records, 0
    lo = max(1, int(ref_length * (1 - _LENGTH_TOLERANCE)))
    hi = int(ref_length * (1 + _LENGTH_TOLERANCE))
    kept = [r for r in records if lo <= len(r.seq) <= hi]
    return kept, len(records) - len(kept)


def _is_non_interactive() -> bool:
    return not sys.stdin.isatty()


# ── Threshold prompt ──────────────────────────────────────────────────────────

def prompt_analysis_threshold(project_name: str, project_config: dict) -> float:
    """
    Returns the analysis threshold for this run.
    Interactive: shows current value and lets the user change it.
    Non-interactive: uses saved value or default 1.0.
    Persists to project_config['conservation_threshold'] after any change.
    """
    saved = project_config.get("conservation_threshold")

    if _is_non_interactive():
        return float(saved) if saved is not None else 1.0

    if saved is not None:
        console.print(Panel(
            f"[bold]Conservation analysis threshold[/bold]\n\n"
            f"Current threshold: [cyan]{int(round(float(saved) * 100))}%[/cyan]\n\n"
            "[dim]Determines which variants are shown as passed/failed.\n"
            "Summary counts (100%, 90%, 80%) and row colours are fixed\n"
            "regardless of this value.[/dim]\n\n"
            "  [cyan][1][/cyan] Keep current value\n"
            "  [cyan][2][/cyan] Change threshold",
            box=box.ROUNDED, title="Setup: analyze_conservation", title_align="left",
        ))
        while True:
            try:
                choice = input("> ").strip()
            except EOFError:
                choice = "1"
            if choice in ("1", ""):
                return float(saved)
            if choice == "2":
                break
            console.print("[dim]Type 1 or 2.[/dim]")

    console.print(Panel(
        "[bold]Conservation analysis threshold[/bold]\n\n"
        "[dim]Variants with identity >= threshold appear as 'passed'.\n"
        "Failed variants show their actual best-matching window and mutations.[/dim]\n\n"
        "  [cyan][1][/cyan] 1.00 — exact match only (default)\n"
        "  [cyan][2][/cyan] 0.90 — 90% identity\n"
        "  [cyan][3][/cyan] 0.80 — 80% identity\n"
        "  [cyan][4][/cyan] Custom value (e.g. 0.75, 0.85, 0.65)",
        box=box.ROUNDED, title="Setup: analyze_conservation", title_align="left",
    ))

    preset_map = {"1": 1.0, "2": 0.90, "3": 0.80}
    while True:
        try:
            choice = input("> ").strip()
        except EOFError:
            choice = "1"
        if choice in preset_map:
            threshold = preset_map[choice]
            break
        if choice == "4":
            while True:
                try:
                    raw = input("Threshold (e.g. 0.75): ").strip().replace(",", ".")
                except EOFError:
                    raw = "1.0"
                try:
                    threshold = float(raw)
                    if 0.0 < threshold <= 1.0:
                        break
                    console.print("[red]Value must be between 0 (exclusive) and 1 (inclusive).[/red]")
                except ValueError:
                    console.print("[red]Invalid value. Use a dot as decimal separator (e.g. 0.75).[/red]")
            break
        console.print("[dim]Invalid option. Type 1, 2, 3 or 4.[/dim]")

    project_config["conservation_threshold"] = threshold
    save_project_config(project_name, project_config)
    console.print(f"[dim]Threshold set to {threshold:.2f}, saved to project_config.[/dim]")
    return threshold


# ── IEDB-style summary table builder ──────────────────────────────────────────

_SUMMARY_COLUMNS = [
    "#",
    "peptide",
    "length",
    "pct_identity_100",
    "pct_identity_90",
    "pct_identity_80",
    "pct_identity_threshold",
    "min_identity",
    "max_identity",
    "avg_identity",
    "conservation_label",
    "n_excellent_match",
    "n_tolerated",
    "variants_exact_match",
    "variants_tolerable",
    "alleles_united",
    "num_alleles_united",
]


def _build_iedb_summary_rows(result_df: pd.DataFrame) -> pd.DataFrame:
    """
    Builds the IEDB-style presentation DataFrame. Tier columns are formatted
    as "XX.XX% (n/N)" strings; min/max/avg identity as "XX.XX%". Every other
    field passes through.
    """
    rows = []
    for idx, row in enumerate(result_df.itertuples(index=False), start=1):
        n_total      = int(getattr(row, "n_variants_used", 0))
        peptide      = getattr(row, COLUMN_PEPTIDE)
        num_alleles  = getattr(row, COLUMN_NUM_ALLELES_UNITED, 0)
        num_alleles  = int(num_alleles) if pd.notna(num_alleles) else 0

        rows.append({
            "#":                      idx,
            "peptide":                peptide,
            "length":                 len(peptide),
            "pct_identity_100":       _format_pct_fraction(int(getattr(row, "n_exact_match",      0)), n_total),
            "pct_identity_90":        _format_pct_fraction(int(getattr(row, "n_identity_90",      0)), n_total),
            "pct_identity_80":        _format_pct_fraction(int(getattr(row, "n_identity_80",      0)), n_total),
            "pct_identity_threshold": _format_pct_fraction(int(getattr(row, "n_passed_threshold", 0)), n_total),
            "min_identity":           f"{float(getattr(row, 'min_max_identity',  0.0)) * 100:.2f}%",
            "max_identity":           f"{float(getattr(row, 'max_max_identity',  0.0)) * 100:.2f}%",
            "avg_identity":           f"{float(getattr(row, 'mean_max_identity', 0.0)) * 100:.2f}%",
            "conservation_label":     getattr(row, "conservation_label", "conservation_unknown"),
            "n_excellent_match":      int(getattr(row, "n_excellent_match", 0)),
            "n_tolerated":            int(getattr(row, "n_tolerated", 0)),
            "variants_exact_match":   getattr(row, "variants_exact_match", "") or "",
            "variants_tolerable":     getattr(row, "variants_tolerable",   "") or "",
            "alleles_united":         getattr(row, COLUMN_ALLELES_UNITED, "") or "",
            "num_alleles_united":     num_alleles,
        })
    return pd.DataFrame(rows, columns=_SUMMARY_COLUMNS)


# ── XLSX writers ──────────────────────────────────────────────────────────────

def write_conservation_summary_xlsx(df: pd.DataFrame, output_path: Path):
    """
    IEDB-style XLSX: header gray, no row backgrounds, only `conservation_label`
    cell coloured. Single sheet, frozen header.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Conservation Summary"

    columns = list(df.columns)

    for col_idx, col_name in enumerate(columns, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = _FILL_GRAY

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            if col_name == "conservation_label":
                cell.fill      = _LABEL_FILL.get(row[col_name], _FILL_UNKNOWN)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "#":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_width  = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].tolist()),
        )
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    wb.save(str(output_path))


_MUTATIONS_COLUMNS = [
    "peptide", "length", "variant_accession", "variant_window",
    "identity", "n_mutations", "mutations", "mutation_positions",
    "anchor_hit", "blosum62_min", "mhc_verdict", "alleles_united",
]


def write_mutations_xlsx(mutation_records: list[dict], output_path: Path):
    """
    Per (epitope, variant) breakdown for variants with 1 or 2 mutations.
    Rows coloured by `mhc_verdict`; sorted excellent → tolerated → likely_lost.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutation Detail"

    for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = _FILL_GRAY

    if not mutation_records:
        ws.cell(row=2, column=1, value="(no variants with 1 or 2 mutations)")
        ws.freeze_panes = "A2"
        wb.save(str(output_path))
        return

    sorted_records = sorted(
        mutation_records,
        key=lambda r: (
            _VERDICT_ORDER.get(r["mhc_verdict"], 99),
            r["peptide"],
            r["n_mutations"],
        ),
    )

    left_aligned = {"alleles_united", "mutations", "peptide", "variant_window", "variant_accession"}
    for row_idx, record in enumerate(sorted_records, start=2):
        row_fill = _VERDICT_FILL.get(record["mhc_verdict"], _FILL_UNKNOWN)
        for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name))
            cell.fill      = row_fill
            cell.alignment = Alignment(
                horizontal="left" if col_name in left_aligned else "center",
                vertical="center",
            )

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(_MUTATIONS_COLUMNS, start=1):
        col_letter      = ws.cell(row=1, column=col_idx).column_letter
        column_values   = [str(r.get(col_name, "")) for r in sorted_records]
        max_width       = max([len(col_name)] + [len(v) for v in column_values])
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)

    wb.save(str(output_path))


# ── Rich console output ───────────────────────────────────────────────────────

def print_conservation_rich_table(
    df: pd.DataFrame, track_id: str, analysis_threshold: float
):
    thr_pct = int(round(analysis_threshold * 100))
    t = Table(
        box=box.SIMPLE,
        title=f"Conservation — {track_id}  (threshold={thr_pct}%)",
        show_lines=False,
    )
    t.add_column("Peptide",         style="bold", no_wrap=True)
    t.add_column("Variants",        justify="right")
    t.add_column(f"≥{thr_pct}%",    justify="right")
    t.add_column("100%",            justify="right")
    t.add_column("≥90%",            justify="right")
    t.add_column("≥80%",            justify="right")
    t.add_column("Avg ID",          justify="right")
    t.add_column("Label",           no_wrap=True)

    label_rich = {
        "perfect":              "[bold green]★ perfect[/bold green]",
        "high":                 "[green]  high[/green]",
        "moderate":             "[yellow]  moderate[/yellow]",
        "low":                  "[red]  low[/red]",
        "conservation_unknown": "[dim]  unknown[/dim]",
    }

    for _, row in df.iterrows():
        label   = row.get("conservation_label", "conservation_unknown")
        n_total = int(row.get("n_variants_used", 0))

        def fmt(num: int) -> str:
            return f"{num}/{n_total}" if n_total else "—"

        t.add_row(
            str(row.get(COLUMN_PEPTIDE, "")),
            str(n_total),
            fmt(int(row.get("n_passed_threshold", 0))),
            fmt(int(row.get("n_exact_match",      0))),
            fmt(int(row.get("n_identity_90",      0))),
            fmt(int(row.get("n_identity_80",      0))),
            f"{float(row.get('mean_max_identity', 0.0)) * 100:.1f}%",
            label_rich.get(label, label),
        )

    console.print(t)


# ── Dual-panel PNG (position conservation + identity tiers) ──────────────────

def write_conservation_dual_panel_png(
    alignment_data: list[dict],
    track_id: str,
    n_variants_total: int,
    analysis_threshold: float,
    output_path: Path,
):
    """
    Single figure with three axes sharing the Y axis (peptides):
      [label sidebar] [position conservation P1..PΩ] [identity tiers]

    Filter: epitopes with ≥1 variant having ≤_MAX_MUTATIONS_FOR_HEATMAP
    substitutions. This includes 100%-conserved epitopes (0 muts ≤ 2).
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
    import seaborn as sns

    filtered = [
        item for item in alignment_data
        if _has_variant_within_max_mut(item["peptide"], item["alignment_tuples"])
    ]
    if not filtered or n_variants_total == 0:
        return

    enriched: list[dict] = []
    for item in filtered:
        peptide              = item["peptide"]
        pos_stats            = _compute_position_stats(peptide, item["alignment_tuples"])
        n_positions          = len(peptide)
        p2_idx, pomega_idx   = _anchor_positions(n_positions)
        p2_pct               = pos_stats[p2_idx]["conservation_pct"]     if len(pos_stats) > p2_idx     else 0.0
        pomega_pct           = pos_stats[pomega_idx]["conservation_pct"] if len(pos_stats) > pomega_idx else 0.0
        enriched.append({
            "peptide":            peptide,
            "pos_stats":          pos_stats,
            "anchor_score":       min(p2_pct, pomega_pct),
            "n_positions":        n_positions,
            "conservation_label": item["conservation_label"],
            "n_exact_match":      item["n_exact_match"],
            "n_identity_90":      item["n_identity_90"],
            "n_identity_80":      item["n_identity_80"],
            "n_passed_threshold": item["n_passed_threshold"],
        })

    enriched.sort(key=lambda x: x["anchor_score"], reverse=True)

    n_rows         = len(enriched)
    n_positions    = max(item["n_positions"] for item in enriched)
    p2_idx, pomega_idx = _anchor_positions(n_positions)
    threshold_pct  = int(round(analysis_threshold * 100))

    # Anchor-aligned matrix: every peptide's P2 sits at col 1 and PΩ at the
    # last col, regardless of its length. The middle bulge becomes NaN.
    pos_matrix   = np.full((n_rows, n_positions), np.nan, dtype=float)
    pos_annot    = np.empty((n_rows, n_positions), dtype=object)
    pos_annot[:] = ""
    for row_idx, item in enumerate(enriched):
        aligned_pcts, aligned_annot = _align_position_stats_to_anchor(
            item["pos_stats"], item["n_positions"], n_positions
        )
        for col_idx, pct in enumerate(aligned_pcts):
            if pct is not None:
                pos_matrix[row_idx, col_idx] = pct
            pos_annot[row_idx, col_idx] = aligned_annot[col_idx]

    tier_matrix = np.array(
        [
            [
                item["n_exact_match"]      / n_variants_total * 100,
                item["n_identity_90"]      / n_variants_total * 100,
                item["n_identity_80"]      / n_variants_total * 100,
                item["n_passed_threshold"] / n_variants_total * 100,
            ]
            for item in enriched
        ]
    )
    tier_annot = np.array(
        [
            [
                f"{item['n_exact_match']}/{n_variants_total}",
                f"{item['n_identity_90']}/{n_variants_total}",
                f"{item['n_identity_80']}/{n_variants_total}",
                f"{item['n_passed_threshold']}/{n_variants_total}",
            ]
            for item in enriched
        ]
    )

    peptides = [item["peptide"]            for item in enriched]
    labels   = [item["conservation_label"] for item in enriched]

    fig_height = max(4.0, min(22.0, 0.45 * n_rows + 2.0))
    fig_width  = 14.0

    fig, axes = plt.subplots(
        1, 3,
        figsize=(fig_width, fig_height),
        gridspec_kw={"width_ratios": [0.04, n_positions * 0.55, 4 * 0.7]},
    )

    # ─── Left: label sidebar ───
    ax_bar = axes[0]
    for i, lbl in enumerate(labels):
        ax_bar.add_patch(
            mpatches.Rectangle((0, i), 1, 1, color=_LABEL_HEX.get(lbl, "#D9D9D9"), linewidth=0)
        )
    ax_bar.set_xlim(0, 1)
    ax_bar.set_ylim(0, n_rows)
    ax_bar.set_xticks([])
    ax_bar.set_yticks(np.arange(n_rows) + 0.5)
    ax_bar.set_yticklabels(peptides, fontsize=8, fontfamily="monospace")
    ax_bar.invert_yaxis()
    ax_bar.set_xlabel("Label", fontsize=8)
    ax_bar.tick_params(axis="y", length=0)

    # ─── Middle: position conservation ───
    ax_pos = axes[1]
    show_annot = n_rows <= 60
    sns.heatmap(
        pos_matrix,
        ax=ax_pos,
        cmap="RdYlGn",
        vmin=0, vmax=100,
        annot=pos_annot if show_annot else False,
        fmt="",
        annot_kws={"size": 7},
        linewidths=0.4, linecolor="#cccccc",
        cbar=False,
        xticklabels=_build_anchor_aligned_xtick_labels(n_positions),
        yticklabels=False,
        mask=np.isnan(pos_matrix),
    )
    ax_pos.set_facecolor("#f0f0f0")
    ax_pos.set_xlabel("")
    ax_pos.set_title("Position conservation", fontsize=10, fontweight="bold", pad=8)
    ax_pos.tick_params(axis="x", labelsize=9)

    for anchor_idx in (p2_idx, pomega_idx):
        ax_pos.add_patch(plt.Rectangle(
            (anchor_idx, 0), 1, n_rows,
            fill=False, edgecolor=_ANCHOR_BORDER_HEX, linewidth=2.5, clip_on=False,
        ))

    # ─── Right: identity tiers ───
    ax_tier = axes[2]
    tier_headers = ["≥100%", "≥90%", "≥80%", f"≥{threshold_pct}%"]
    sns.heatmap(
        tier_matrix,
        ax=ax_tier,
        cmap="RdYlGn",
        vmin=0, vmax=100,
        annot=tier_annot if show_annot else False,
        fmt="",
        annot_kws={"size": 7},
        linewidths=0.4, linecolor="#cccccc",
        cbar_kws={"shrink": 0.7, "label": "% of variants"},
        xticklabels=tier_headers,
        yticklabels=False,
    )
    ax_tier.set_xlabel("")
    ax_tier.set_title(f"Identity tiers (n={n_variants_total} variants)", fontsize=10, fontweight="bold", pad=8)
    ax_tier.tick_params(axis="x", labelsize=9)

    fig.suptitle(
        f"Conservation — {track_id} ({n_rows} epitopes shown, {n_variants_total} variants analysed)",
        fontsize=12, fontweight="bold", y=0.995,
    )

    legend_patches = [
        mpatches.Patch(color=_LABEL_HEX["perfect"],  label="★ Perfect"),
        mpatches.Patch(color=_LABEL_HEX["high"],     label="High"),
        mpatches.Patch(color=_LABEL_HEX["moderate"], label="Moderate"),
        mpatches.Patch(color=_LABEL_HEX["low"],      label="Low"),
        mpatches.Patch(facecolor="white", edgecolor=_ANCHOR_BORDER_HEX, linewidth=2, label="Anchor (P2, PΩ)"),
    ]
    fig.legend(
        handles=legend_patches, loc="lower center", ncol=5,
        fontsize=8, frameon=False, bbox_to_anchor=(0.5, 0.005),
    )

    plt.tight_layout(rect=[0, 0.03, 1, 0.97])
    fig.savefig(str(output_path), dpi=150, bbox_inches="tight")
    plt.close(fig)


# ── Step infrastructure: preflight / postflight helpers ───────────────────────

def _inspect_track_fasta_status(
    project_name: str,
    track_id: str,
) -> dict:
    """
    Return a status dict describing the variant FASTA situation for one track.

    Used by `preflight()` to render the consolidated status table before the
    user is asked whether they want to provide local FASTA files. The dict
    has the shape:

        {
            "track_id":      str,
            "ref_length":    int | None,
            "fasta_path":    Path | None,
            "fasta_exists":  bool,
            "n_records_raw": int,
            "n_records_kept": int,
            "status":        str,   # "ok" | "missing_fasta" | "no_reference" | "length_mismatch"
        }
    """
    project_root = Path("projects") / project_name
    track_dir = project_root / "data" / "intermediate" / track_id
    input_track_dir = project_root / "data" / "input" / track_id

    reference_fasta_path = input_track_dir / get_step_filename("SEQUENCES", track_id, ext="fasta")
    reference_length: Optional[int] = None
    if reference_fasta_path.exists():
        reference_records = list(SeqIO.parse(str(reference_fasta_path), "fasta"))
        if reference_records:
            reference_length = len(reference_records[0].seq)

    variants_fasta_path = track_dir / "variants" / get_step_filename("VARIANTS", track_id, ext="fasta")
    fasta_exists = variants_fasta_path.exists() and variants_fasta_path.stat().st_size > 0

    n_records_raw = 0
    n_records_kept = 0
    if fasta_exists:
        kept_records, excluded_count = load_fasta_sequences(variants_fasta_path, reference_length or 0)
        n_records_kept = len(kept_records)
        n_records_raw = n_records_kept + excluded_count

    if not reference_length:
        status_label = "no_reference"
    elif not fasta_exists:
        status_label = "missing_fasta"
    elif n_records_kept == 0 and n_records_raw > 0:
        status_label = "length_mismatch"
    else:
        status_label = "ok"

    return {
        "track_id":       track_id,
        "ref_length":     reference_length,
        "fasta_path":     variants_fasta_path if fasta_exists else None,
        "fasta_exists":   fasta_exists,
        "n_records_raw":  n_records_raw,
        "n_records_kept": n_records_kept,
        "status":         status_label,
    }


def _render_track_status_table(track_status_rows: list[dict]) -> Table:
    """Build the consolidated Rich Table shown by `preflight()`."""
    status_table = Table(box=box.SIMPLE, header_style="bold cyan")
    status_table.add_column("Track", style="cyan")
    status_table.add_column("Ref length", justify="right")
    status_table.add_column("Variants file", overflow="fold")
    status_table.add_column("n total", justify="right")
    status_table.add_column("n after ±25% filter", justify="right")
    status_table.add_column("Status")

    status_color_map = {
        "ok":              "[green]ok[/green]",
        "missing_fasta":   "[yellow]missing variants FASTA[/yellow]",
        "no_reference":    "[red]no reference FASTA[/red]",
        "length_mismatch": "[yellow]all variants filtered out (length)[/yellow]",
    }

    for status_row in track_status_rows:
        status_table.add_row(
            status_row["track_id"],
            str(status_row["ref_length"]) if status_row["ref_length"] is not None else "—",
            status_row["fasta_path"].name if status_row["fasta_path"] else "—",
            str(status_row["n_records_raw"])  if status_row["fasta_exists"] else "—",
            str(status_row["n_records_kept"]) if status_row["fasta_exists"] else "—",
            status_color_map.get(status_row["status"], status_row["status"]),
        )

    return status_table


def _ask_for_local_fasta_overrides(track_status_rows: list[dict]) -> dict:
    """
    Loop interactively letting the user attach a local FASTA path to one or
    more track ids. Returns {track_id: path_string}. Empty dict if the user
    declines or supplies nothing usable.
    """
    fasta_overrides: dict[str, str] = {}
    track_id_set = {row["track_id"] for row in track_status_rows}

    while True:
        try:
            track_input = input("  Track id to attach FASTA to (or blank to finish): ").strip()
        except EOFError:
            track_input = ""

        if not track_input:
            break
        if track_input not in track_id_set:
            console.print(f"  [yellow]Unknown track id '{track_input}'. Pick one from the table above.[/yellow]")
            continue

        try:
            path_input = input(f"  Local FASTA path for {track_input}: ").strip()
        except EOFError:
            path_input = ""

        candidate_path = Path(path_input).expanduser() if path_input else None
        if candidate_path is None or not candidate_path.exists() or candidate_path.stat().st_size == 0:
            console.print("  [red]File not found or empty. Skipping.[/red]")
            continue

        fasta_overrides[track_input] = str(candidate_path)
        console.print(f"  [green]✓ Will use {candidate_path} for {track_input}.[/green]")

    return fasta_overrides


def _identify_problematic_tracks(
    project_name: str,
    track_outcomes: dict,
    minimum_acceptable_variants: int = 5,
) -> list[dict]:
    """
    Return tracks whose conservation analysis did not produce a meaningful
    result. A track is problematic when (a) it errored, (b) it ran but had
    fewer than `minimum_acceptable_variants` usable sequences, or (c) it ran
    with no variants at all (conservation_unknown).
    """
    problematic_track_rows: list[dict] = []

    for track_id, outcome_dict in track_outcomes.items():
        outcome_status = outcome_dict.get("status")

        if outcome_status == "error":
            problematic_track_rows.append({
                "track_id": track_id,
                "problem":  outcome_dict.get("error_message", "unknown error"),
                "n_used":   None,
            })
            continue

        audit_path = (
            Path("projects") / project_name / "data" / "intermediate" / track_id /
            "conservation" / get_step_filename("CONSERVATION_AUDIT", track_id, ext="json")
        )
        if not audit_path.exists():
            continue

        try:
            audit_payload = json.loads(audit_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        n_variants_used = int(audit_payload.get("n_variants_used", 0))
        if n_variants_used < minimum_acceptable_variants:
            problematic_track_rows.append({
                "track_id": track_id,
                "problem":  f"only {n_variants_used} variants used"
                            f" (label_counts={audit_payload.get('label_counts', {})})",
                "n_used":   n_variants_used,
            })

    return problematic_track_rows


# ── Step ──────────────────────────────────────────────────────────────────────

class AnalyzeConservationStep(BaseTrackStep):
    step_name   = "analyze_conservation"
    description = (
        "Slides every ★ epitope across each variant sequence, records best-match "
        "identity, and labels each peptide as perfect / high / moderate / low. "
        "Also emits per-variant mutation verdicts (BLOSUM62 + MHC-I anchors P2 "
        "and PΩ). Qualitative — never removes epitopes."
    )

    @classmethod
    def preflight(cls, project_name: str, project_config: dict, track_ids: list[str]) -> Optional[dict]:
        if not track_ids:
            return None

        track_status_rows = [
            _inspect_track_fasta_status(project_name, track_id)
            for track_id in track_ids
        ]

        console.print(Panel(
            _render_track_status_table(track_status_rows),
            title="[bold]Variant FASTA status across all tracks[/bold]",
            border_style="cyan", box=box.ROUNDED,
        ))

        if not is_interactive_session():
            return None

        try:
            response = input(
                "  Do you have a local FASTA you want to attach to any track? [y/N]: "
            ).strip().lower()
        except EOFError:
            response = ""

        if response not in {"y", "yes"}:
            return None

        console.print(
            "[dim]  Type the track id (must match the table above), then the FASTA path. "
            "Leave the track id blank to stop.[/dim]"
        )
        fasta_overrides = _ask_for_local_fasta_overrides(track_status_rows)
        return fasta_overrides if fasta_overrides else None

    @classmethod
    def postflight(cls, project_name: str, project_config: dict, track_outcomes: dict) -> None:
        if not track_outcomes or not is_interactive_session():
            return

        problematic_track_rows = _identify_problematic_tracks(project_name, track_outcomes)
        if not problematic_track_rows:
            return

        recovery_table = Table(box=box.SIMPLE, header_style="bold yellow")
        recovery_table.add_column("Track", style="cyan")
        recovery_table.add_column("Variants used", justify="right")
        recovery_table.add_column("Problem")
        for problem_row in problematic_track_rows:
            recovery_table.add_row(
                problem_row["track_id"],
                "—" if problem_row["n_used"] is None else str(problem_row["n_used"]),
                problem_row["problem"],
            )

        console.print(Panel(
            recovery_table,
            title="[bold]Tracks with weak conservation results[/bold]",
            border_style="yellow", box=box.ROUNDED,
        ))

        try:
            response = input(
                "  Re-run any of these with a local FASTA? [y/N]: "
            ).strip().lower()
        except EOFError:
            response = ""
        if response not in {"y", "yes"}:
            return

        from utils.pipeline_state import reset_track_step

        problematic_track_id_set = {row["track_id"] for row in problematic_track_rows}
        while True:
            try:
                target_track = input("  Track id to re-run (or blank to finish): ").strip()
            except EOFError:
                target_track = ""
            if not target_track:
                break
            if target_track not in problematic_track_id_set:
                console.print(f"  [yellow]Not in the list above: '{target_track}'.[/yellow]")
                continue

            try:
                fasta_path_input = input(f"  Local FASTA path for {target_track}: ").strip()
            except EOFError:
                fasta_path_input = ""
            override_path = Path(fasta_path_input).expanduser() if fasta_path_input else None
            if override_path is None or not override_path.exists() or override_path.stat().st_size == 0:
                console.print("  [red]File not found or empty. Skipping.[/red]")
                continue

            reset_track_step(project_name, target_track, cls.step_name)
            rerun_step_instance = cls(
                project_name=project_name,
                project_config=project_config,
                track_id=target_track,
            )
            rerun_step_instance.preflight_config = {target_track: str(override_path)}
            rerun_step_instance.execute(force_rerun=False)

    def describe_outputs(self) -> dict[Path, str]:
        conservation_dir = self.track_dir / "conservation"
        return {
            conservation_dir / get_step_filename("CONSERVATION", self.track_id):
                "IEDB-style summary: per ★ rep with tier %, fractions, min/max/avg identity, label.",
            conservation_dir / get_step_filename("CONSERVATION", self.track_id, ext="xlsx"):
                "Same table with header-only styling; conservation_label cell coloured.",
            conservation_dir / get_step_filename("CONSERVATION_HEATMAP", self.track_id, ext="png"):
                "Dual-panel heatmap: position conservation + identity tiers (filtered to ≤2-mut variants).",
            conservation_dir / get_step_filename("CONSERVATION_MUTATIONS", self.track_id, ext="xlsx"):
                "Per (epitope, variant) breakdown with anchor flag, BLOSUM62 score, MHC verdict.",
            conservation_dir / get_step_filename("CONSERVATION_AUDIT", self.track_id, ext="json"):
                "Run audit — threshold, FASTA source, variants used, label counts, verdict counts.",
        }

    def run(self, input_data=None):
        analysis_threshold = prompt_analysis_threshold(
            self.project_name, self.project_config
        )

        # ── Load ★ representatives ────────────────────────────────────────────
        clusters_dir        = self.track_dir / "clusters"
        representatives_csv = clusters_dir / get_step_filename("CLUSTER_REPR", self.track_id)
        if not representatives_csv.exists():
            raise FileNotFoundError(
                f"select_representatives output not found: {representatives_csv}\n"
                "Run 'select_representatives' before 'analyze_conservation'."
            )

        df_repr  = pd.read_csv(representatives_csv)
        df_stars = df_repr[df_repr[COLUMN_BEST_REPRESENTATIVE] == STAR_MARKER].copy()
        if df_stars.empty:
            raise ValueError(
                f"No ★ representatives found in {representatives_csv.name}. "
                "Run 'select_representatives' first."
            )

        # ── Reference length (for variant length filtering) ───────────────────
        input_dir  = self.track_dir.parent.parent / "input" / self.track_id
        ref_fasta  = input_dir / get_step_filename("SEQUENCES", self.track_id, ext="fasta")
        ref_length = 0
        if ref_fasta.exists():
            ref_recs = list(SeqIO.parse(str(ref_fasta), "fasta"))
            if ref_recs:
                ref_length = len(ref_recs[0].seq)
                console.print(
                    f"[dim]→ Reference length: {ref_length} aa "
                    f"(±{int(_LENGTH_TOLERANCE*100)}% filter applied)[/dim]"
                )

        # ── Resolve FASTA source ──────────────────────────────────────────────
        variants_dir = self.track_dir / "variants"
        fasta_path   = variants_dir / get_step_filename("VARIANTS", self.track_id, ext="fasta")
        fasta_source = "none"
        records: list = []

        override_fasta_path: Optional[Path] = None
        if isinstance(self.preflight_config, dict):
            override_value = self.preflight_config.get(self.track_id)
            if override_value:
                override_fasta_path = Path(override_value).expanduser()

        if override_fasta_path is not None and override_fasta_path.exists() and override_fasta_path.stat().st_size > 0:
            records, n_excl = load_fasta_sequences(override_fasta_path, ref_length)
            fasta_source = "user_provided"
            msg = f"[dim]→ Using user-provided FASTA {override_fasta_path.name} ({len(records)} sequences"
            if n_excl:
                msg += f", {n_excl} excluded by length filter"
            console.print(msg + ")[/dim]")
        elif fasta_path.exists() and fasta_path.stat().st_size > 0:
            records, n_excl = load_fasta_sequences(fasta_path, ref_length)
            fasta_source    = "variants_cache"
            msg = f"[dim]→ Using {fasta_path.name} ({len(records)} sequences"
            if n_excl:
                msg += f", {n_excl} excluded by length filter"
            console.print(msg + ")[/dim]")
        else:
            console.print(
                f"[yellow]⚠ No variant FASTA available for {self.track_id} — "
                "epitopes will be labelled 'conservation_unknown'. Provide a "
                "local FASTA after the loop completes if needed.[/yellow]"
            )

        # ── Analyse each peptide ──────────────────────────────────────────────
        metrics_rows:        list[dict] = []
        alignment_data:      list[dict] = []
        all_mutation_records: list[dict] = []

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            MofNCompleteColumn(),
            TimeElapsedColumn(),
            console=console,
            transient=True,
        ) as conservation_progress_bar:
            conservation_task_id = conservation_progress_bar.add_task(
                "  Analyzing per-epitope conservation",
                total=len(df_stars),
            )
            for _, repr_row in df_stars.iterrows():
                peptide = repr_row[COLUMN_PEPTIDE]
                alleles_united_value = repr_row.get(COLUMN_ALLELES_UNITED, "")
                alleles_united_str   = str(alleles_united_value) if pd.notna(alleles_united_value) else ""

                metrics, alignment_tuples = compute_epitope_conservation(
                    peptide, records, analysis_threshold
                )

                peptide_mutation_records = compute_mutation_records(
                    peptide, alignment_tuples, alleles_united_str
                )
                all_mutation_records.extend(peptide_mutation_records)

                variants_exact_labels = [
                    label for label, identity, _window in alignment_tuples
                    if identity == 1.0
                ]
                tolerable_records = [
                    r for r in peptide_mutation_records
                    if r["mhc_verdict"] in {"excellent_match", "tolerated"}
                ]
                n_excellent = sum(1 for r in peptide_mutation_records if r["mhc_verdict"] == "excellent_match")
                n_tolerated = sum(1 for r in peptide_mutation_records if r["mhc_verdict"] == "tolerated")

                metrics["n_excellent_match"]    = n_excellent
                metrics["n_tolerated"]          = n_tolerated
                metrics["variants_exact_match"] = "; ".join(variants_exact_labels)
                metrics["variants_tolerable"]   = "; ".join(
                    f"{r['variant_accession']}[{r['mutations']}]" for r in tolerable_records
                )

                metrics_rows.append(metrics)

                alignment_data.append({
                    "peptide":            peptide,
                    "alignment_tuples":   alignment_tuples,
                    "conservation_label": metrics["conservation_label"],
                    "n_exact_match":      metrics["n_exact_match"],
                    "n_identity_90":      metrics["n_identity_90"],
                    "n_identity_80":      metrics["n_identity_80"],
                    "n_passed_threshold": metrics["n_passed_threshold"],
                })

                conservation_progress_bar.advance(conservation_task_id)

        # ── Build full result DataFrame (metrics + repr columns) ──────────────
        result_df = pd.concat(
            [df_stars.reset_index(drop=True), pd.DataFrame(metrics_rows)],
            axis=1,
        )

        # ── Save outputs ──────────────────────────────────────────────────────
        conservation_dir = self.track_dir / "conservation"
        conservation_dir.mkdir(parents=True, exist_ok=True)

        output_csv             = conservation_dir / get_step_filename("CONSERVATION",            self.track_id)
        output_xlsx            = conservation_dir / get_step_filename("CONSERVATION",            self.track_id, ext="xlsx")
        output_heatmap_png     = conservation_dir / get_step_filename("CONSERVATION_HEATMAP",    self.track_id, ext="png")
        output_mutations_xlsx  = conservation_dir / get_step_filename("CONSERVATION_MUTATIONS",  self.track_id, ext="xlsx")
        audit_path             = conservation_dir / get_step_filename("CONSERVATION_AUDIT",      self.track_id, ext="json")

        # Drop deprecated files if present from previous runs
        for deprecated_name in ("CONSERVATION_VISUAL", "CONSERVATION_POSITIONS"):
            deprecated_path = conservation_dir / get_step_filename(deprecated_name, self.track_id, ext="xlsx")
            if deprecated_path.exists():
                deprecated_path.unlink()
            deprecated_png_path = conservation_dir / get_step_filename(deprecated_name, self.track_id, ext="png")
            if deprecated_png_path.exists():
                deprecated_png_path.unlink()

        iedb_summary_df = _build_iedb_summary_rows(result_df)
        iedb_summary_df.to_csv(output_csv, index=False)
        write_conservation_summary_xlsx(iedb_summary_df, output_xlsx)
        write_mutations_xlsx(all_mutation_records, output_mutations_xlsx)
        write_conservation_dual_panel_png(
            alignment_data, self.track_id, len(records), analysis_threshold, output_heatmap_png,
        )

        # ── Console summary ───────────────────────────────────────────────────
        print_conservation_rich_table(result_df, self.track_id, analysis_threshold)

        n_analyzed   = len(result_df)
        label_counts = result_df["conservation_label"].value_counts().to_dict()
        mean_all     = round(float(result_df["mean_max_identity"].mean()), 4) if n_analyzed else 0.0

        verdict_counts = dict(Counter(r["mhc_verdict"] for r in all_mutation_records))
        n_epitopes_with_tolerant = len({r["peptide"] for r in all_mutation_records})

        console.print(
            f"\n[bold green]RESULT: {n_analyzed} representatives analysed "
            f"({fasta_source}, threshold={int(round(analysis_threshold * 100))}%).[/bold green]\n"
        )

        # ── Audit JSON ────────────────────────────────────────────────────────
        audit = {
            "timestamp":                         datetime.datetime.now().isoformat(),
            "track_id":                          self.track_id,
            "analysis_threshold":                analysis_threshold,
            "fasta_source":                      fasta_source,
            "n_variants_used":                   len(records),
            "ref_length_aa":                     ref_length,
            "length_filter_tolerance":           _LENGTH_TOLERANCE,
            "n_representatives_analyzed":        n_analyzed,
            "label_counts":                      label_counts,
            "mean_identity_across_all_epitopes": mean_all,
            "n_pairs_within_2mut":               len(all_mutation_records),
            "verdict_counts":                    verdict_counts,
            "n_epitopes_with_tolerant_variants": n_epitopes_with_tolerant,
            "substitution_criterion":            "BLOSUM62 >= 0",
            "anchor_positions":                  "P2 + PΩ (universal HLA-I)",
            "reference_thresholds": {
                "high":     config.CONSERVATION_HIGH_THRESHOLD,
                "moderate": config.CONSERVATION_MODERATE_THRESHOLD,
            },
            "outputs": {
                "csv":            str(output_csv),
                "summary_xlsx":   str(output_xlsx),
                "heatmap_png":    str(output_heatmap_png),
                "mutations_xlsx": str(output_mutations_xlsx),
                "audit":          str(audit_path),
            },
        }
        audit_path.write_text(json.dumps(audit, indent=2, ensure_ascii=False))

        return {
            "output_csv":            str(output_csv),
            "output_xlsx":           str(output_xlsx),
            "output_heatmap_png":    str(output_heatmap_png),
            "output_mutations_xlsx": str(output_mutations_xlsx),
            "n_analyzed":            n_analyzed,
            "fasta_source":          fasta_source,
            "label_counts":          label_counts,
        }
