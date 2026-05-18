"""cluster_epitopes step.

Groups safe epitopes into sequence-similarity clusters with NetworkX.
Methods: cluster_break (default), single_linkage, clique — see the step class
methodology field for the full description.

Input  (first found): toxicity/TOXICITY_SAFE_{track_id}.csv
                   or consensus/CONSENSUS_IMMUNOGENIC_{track_id}.csv
Output: clusters/CLUSTER_{track_id}.{csv,xlsx} + CLUSTER_AUDIT_{track_id}.json
"""

import json
import datetime
from pathlib import Path

import networkx as nx
import numpy as np
import pandas as pd
from Bio.Align import PairwiseAligner
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rich import box
from utils.console import console, flush_stdin
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, BarColumn, MofNCompleteColumn, TextColumn
from rich.table import Table

import config
from modules.base_step import BaseTrackStep
from utils.naming import get_step_filename, COLUMN_PEPTIDE
from utils.project_manager import save_project_config

_CANONICAL_AMINO_ACIDS = set("ACDEFGHIKLMNPQRSTVWY")


# ── Input resolution ──────────────────────────────────────────────────────────

def _find_input_file(track_dir: Path, track_id: str) -> Path:
    """Returns the best available input file for clustering, in priority order."""
    candidate_paths = [
        track_dir / "toxicity"  / get_step_filename("TOXICITY_SAFE",        track_id),
        track_dir / "consensus" / get_step_filename("CONSENSUS_IMMUNOGENIC", track_id),
    ]
    for candidate_path in candidate_paths:
        if candidate_path.exists():
            return candidate_path
    raise FileNotFoundError(
        "No input file found for cluster_epitopes. Tried:\n"
        + "\n".join(f"  {candidate_path}" for candidate_path in candidate_paths)
    )


# ── Pairwise similarity ───────────────────────────────────────────────────────

def _build_similarity_matrix(epitope_sequences: list) -> np.ndarray:
    """
    Computes the pairwise sequence identity matrix for all epitope pairs.

    Identity = alignment score / max(len_a, len_b), using global alignment with
    match=1, mismatch=0, no gap penalties. This matches the IEDB clustering
    tool's percent-identity definition and produces a value in [0.0, 1.0].
    """
    sequence_aligner = PairwiseAligner()
    sequence_aligner.mode          = "global"
    sequence_aligner.match_score   = 1
    sequence_aligner.mismatch_score = 0
    sequence_aligner.open_gap_score = 0
    sequence_aligner.extend_gap_score = 0

    number_of_sequences = len(epitope_sequences)
    similarity_matrix = np.zeros((number_of_sequences, number_of_sequences), dtype=np.float32)
    np.fill_diagonal(similarity_matrix, 1.0)

    total_pairs = number_of_sequences * (number_of_sequences - 1) // 2

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(
            style="yellow",
            complete_style="green",
            finished_style="green",
            pulse_style="bold yellow",
        ),
        MofNCompleteColumn(),
        console=console,
        transient=True,
    ) as progress_bar:
        similarity_task = progress_bar.add_task(
            "  Computing pairwise sequence identity...", total=total_pairs
        )
        for row_index in range(number_of_sequences):
            for col_index in range(row_index + 1, number_of_sequences):
                alignment_score = sequence_aligner.score(
                    epitope_sequences[row_index], epitope_sequences[col_index]
                )
                max_sequence_length = max(
                    len(epitope_sequences[row_index]),
                    len(epitope_sequences[col_index])
                )
                identity = float(alignment_score) / max_sequence_length if max_sequence_length else 0.0
                similarity_matrix[row_index, col_index] = identity
                similarity_matrix[col_index, row_index] = identity
                progress_bar.advance(similarity_task)

    flush_stdin()

    return similarity_matrix


# ── XLSX writer with alternating cluster bands ────────────────────────────────

_XLSX_FILL_HEADER  = PatternFill("solid", fgColor="D3D3D3")
_XLSX_FILL_BAND_A  = PatternFill("solid", fgColor="FFFFFF")
_XLSX_FILL_BAND_B  = PatternFill("solid", fgColor="EAF3FF")
_XLSX_FILL_CLUSTER = PatternFill("solid", fgColor="FFE4B5")
_XLSX_FILL_SIZE    = PatternFill("solid", fgColor="FFB6C1")
_XLSX_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)


def _write_cluster_xlsx(epitopes_dataframe: pd.DataFrame, output_path: Path) -> None:
    """Cluster table sorted by cluster_id with alternating bands per cluster.
    Header gray bold; body white/light-blue per cluster_id; cluster_id orange, cluster_size pink."""
    sort_columns = [c for c in ("cluster_id", COLUMN_PEPTIDE) if c in epitopes_dataframe.columns]
    if sort_columns:
        df = epitopes_dataframe.sort_values(sort_columns).reset_index(drop=True)
    else:
        df = epitopes_dataframe.reset_index(drop=True)

    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clusters"
    columns = list(df.columns)

    header_font = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.alignment = center
        cell.fill      = _XLSX_FILL_HEADER
        cell.border    = _XLSX_BORDER

    band_toggle = 0
    previous_cluster_id = None
    for row_index, df_row in df.iterrows():
        cluster_id_value = df_row.get("cluster_id")
        if cluster_id_value != previous_cluster_id:
            band_toggle = 1 - band_toggle
            previous_cluster_id = cluster_id_value
        band_fill = _XLSX_FILL_BAND_A if band_toggle == 0 else _XLSX_FILL_BAND_B

        for col_idx, col_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index + 2, column=col_idx, value=df_row[col_name])
            cell.border = _XLSX_BORDER
            if col_name == "cluster_id":
                cell.fill = _XLSX_FILL_CLUSTER
                cell.font = header_font
            elif col_name == "cluster_size":
                cell.fill = _XLSX_FILL_SIZE
            else:
                cell.fill = band_fill

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        cell_widths = [len(str(v)) for v in df[col_name].tolist()]
        max_width = max([len(str(col_name))] + cell_widths)
        worksheet.column_dimensions[col_letter].width = min(max_width + 2, 60)

    worksheet.freeze_panes = "A2"
    workbook.save(str(output_path))


# ── Graph construction ────────────────────────────────────────────────────────

def _build_similarity_graph(similarity_matrix: np.ndarray, identity_threshold: float) -> nx.Graph:
    """Threshold graph: weighted edge i→j iff similarity[i,j] >= identity_threshold."""
    number_of_nodes = similarity_matrix.shape[0]
    similarity_graph = nx.Graph()
    similarity_graph.add_nodes_from(range(number_of_nodes))
    for row_index in range(number_of_nodes):
        for col_index in range(row_index + 1, number_of_nodes):
            if similarity_matrix[row_index, col_index] >= identity_threshold:
                similarity_graph.add_edge(
                    row_index, col_index,
                    weight=float(similarity_matrix[row_index, col_index])
                )
    return similarity_graph


# ── Clustering algorithms ─────────────────────────────────────────────────────

def _run_single_linkage(similarity_matrix: np.ndarray, identity_threshold: float) -> np.ndarray:
    """Single-linkage: connected components of the threshold graph."""
    similarity_graph = _build_similarity_graph(similarity_matrix, identity_threshold)
    cluster_labels = np.empty(similarity_matrix.shape[0], dtype=int)
    for cluster_id, connected_component in enumerate(nx.connected_components(similarity_graph)):
        for node_index in connected_component:
            cluster_labels[node_index] = cluster_id
    return cluster_labels


def _run_clique_clustering(similarity_matrix: np.ndarray, identity_threshold: float) -> np.ndarray:
    """Clique clustering: assign each peptide to its largest maximal clique (greedy, size-desc).
    Overlaps go to the larger/earlier clique; isolated nodes become singletons."""
    similarity_graph = _build_similarity_graph(similarity_matrix, identity_threshold)
    maximal_cliques  = sorted(nx.find_cliques(similarity_graph), key=len, reverse=True)

    number_of_nodes = similarity_matrix.shape[0]
    cluster_labels  = np.full(number_of_nodes, -1, dtype=int)
    next_cluster_id = 0

    for clique_members in maximal_cliques:
        unassigned_members = [node for node in clique_members if cluster_labels[node] == -1]
        if unassigned_members:
            for node in unassigned_members:
                cluster_labels[node] = next_cluster_id
            next_cluster_id += 1

    # Remaining unassigned nodes (isolated — not in any edge) become singletons
    for node_index in range(number_of_nodes):
        if cluster_labels[node_index] == -1:
            cluster_labels[node_index] = next_cluster_id
            next_cluster_id += 1

    return cluster_labels


def _compute_average_intra_cluster_similarity(member_indices: list, similarity_matrix: np.ndarray) -> float:
    """Mean of every pairwise similarity inside the cluster (including pairs without an edge)."""
    number_of_members = len(member_indices)
    if number_of_members <= 1:
        return 1.0
    total_similarity = sum(
        similarity_matrix[member_indices[i], member_indices[j]]
        for i in range(number_of_members)
        for j in range(i + 1, number_of_members)
    )
    number_of_pairs = number_of_members * (number_of_members - 1) / 2
    return float(total_similarity) / number_of_pairs


def _run_cluster_break(similarity_matrix: np.ndarray, identity_threshold: float) -> np.ndarray:
    """Cluster-break: connected components, then iteratively drop the weakest edge from any
    component whose mean intra-cluster similarity falls below identity_threshold."""
    similarity_graph = _build_similarity_graph(similarity_matrix, identity_threshold)

    graph_modified = True
    while graph_modified:
        graph_modified = False
        for connected_component in list(nx.connected_components(similarity_graph)):
            if len(connected_component) <= 1:
                continue
            component_member_indices = list(connected_component)
            average_similarity = _compute_average_intra_cluster_similarity(
                component_member_indices, similarity_matrix
            )
            if average_similarity < identity_threshold:
                component_edges = [
                    (node_a, node_b, similarity_graph[node_a][node_b]["weight"])
                    for node_a, node_b in similarity_graph.edges(connected_component)
                ]
                if component_edges:
                    weakest_node_a, weakest_node_b, _ = min(
                        component_edges, key=lambda edge_tuple: edge_tuple[2]
                    )
                    similarity_graph.remove_edge(weakest_node_a, weakest_node_b)
                    graph_modified = True
                    break

    cluster_labels = np.empty(similarity_matrix.shape[0], dtype=int)
    for cluster_id, connected_component in enumerate(nx.connected_components(similarity_graph)):
        for node_index in connected_component:
            cluster_labels[node_index] = cluster_id
    return cluster_labels


# ── Parameter prompt ──────────────────────────────────────────────────────────

def _ask_clustering_params(project_name: str, project_config: dict) -> tuple:
    """Returns (identity_threshold, clustering_method), prompting once and caching on project_config."""
    saved_threshold = project_config.get("cluster_threshold")
    saved_method    = project_config.get("cluster_method")
    if saved_threshold is not None and saved_method is not None:
        console.print(
            f"[dim]  Clustering params (saved): "
            f"method={saved_method}, threshold={saved_threshold}[/dim]"
        )
        return float(saved_threshold), str(saved_method)

    default_threshold = config.CLUSTER_IDENTITY_CUTOFF

    console.print(Panel(
        "[bold]Step 1 of 2 — Identity threshold[/bold]\n\n"
        "[dim]Epitopes with pairwise sequence identity ≥ threshold are grouped "
        "into the same cluster. One representative is selected per cluster.[/dim]",
        box=box.ROUNDED, title="Setup: cluster_epitopes", title_align="left",
    ))

    while True:
        try:
            raw_threshold = input(
                f"  Identity threshold (0.0–1.0) [{default_threshold}]: "
            ).strip()
            identity_threshold = float(raw_threshold) if raw_threshold else default_threshold
            if 0.0 < identity_threshold <= 1.0:
                break
            console.print("[yellow]  Must be > 0 and ≤ 1.[/yellow]")
        except (ValueError, EOFError):
            identity_threshold = default_threshold
            break

    # Method panel rendered AFTER threshold is confirmed — bundling them confused users.
    console.print(Panel(
        "[bold]Step 2 of 2 — Clustering method[/bold]\n\n"
        "  [cyan][1][/cyan] cluster_break   — cohesive clusters, no bridge groupings [bold](recommended)[/bold]\n"
        "  [cyan][2][/cyan] single_linkage  — most permissive (IEDB-compatible)\n"
        "  [cyan][3][/cyan] clique          — all members pairwise similar (strictest)",
        box=box.ROUNDED, title="Setup: cluster_epitopes", title_align="left",
    ))

    method_choices = {"1": "cluster_break", "2": "single_linkage", "3": "clique"}
    try:
        raw_method = input("  Method [1]: ").strip()
        clustering_method = method_choices.get(raw_method, "cluster_break")
    except EOFError:
        clustering_method = "cluster_break"

    project_config["cluster_threshold"] = identity_threshold
    project_config["cluster_method"]    = clustering_method
    save_project_config(project_name, project_config)

    return identity_threshold, clustering_method


# ── Step ──────────────────────────────────────────────────────────────────────

class ClusterEpitopesStep(BaseTrackStep):
    step_name   = "cluster_epitopes"
    description = (
        "Builds a pairwise similarity graph over the remaining peptides "
        "(NetworkX) and partitions them into clusters at the configured "
        "identity cutoff so each cluster can be represented by a single ★ "
        "epitope downstream."
    )
    long_description = (
        "Two peptides that share ≥ 80% identity bind very similar TCRs and "
        "are immunologically redundant. This step groups them so the next "
        "step (select_representatives) can pick one ★ per cluster instead "
        "of carrying duplicates through the rest of the pipeline.\n\n"
        "Three clustering algorithms are available — all build the same "
        "Biopython pairwise-alignment similarity graph, but differ in how "
        "they decide who belongs to which cluster."
    )
    methodology = (
        "Similarity matrix: Biopython PairwiseAligner global alignment, "
        "identity score (match=1, mismatch=0, gap=0).\n\n"
        "Then three method choices:\n"
        "  • [bold]cluster_break[/bold] (DEFAULT) — connected components of the "
        "thresholded graph, then iteratively prune the weakest edge from any "
        "component whose mean intra-cluster similarity drops below the "
        "cutoff. Biologically motivated: avoids 'bridge' groupings.\n"
        "  • [bold]single_linkage[/bold] — pure connected components. Most "
        "permissive (A and C cluster together if both link to B).\n"
        "  • [bold]clique[/bold] — every peptide assigned to its largest "
        "maximal clique. Most restrictive (members are pairwise similar)."
    )
    references = [
        {
            'authors': 'Cock PJ, Antao T, Chang JT, et al.',
            'title':   'Biopython: freely available Python tools for computational molecular biology and bioinformatics',
            'journal': 'Bioinformatics',
            'year':    2009,
            'doi':     '10.1093/bioinformatics/btp163',
        },
        {
            'authors': 'Bron C, Kerbosch J.',
            'title':   'Algorithm 457: Finding all cliques of an undirected graph',
            'journal': 'Communications of the ACM',
            'year':    1973,
            'doi':     '10.1145/362342.362367',
        },
    ]
    data_format = (
        "Input is automatic — uses TOXICITY_SAFE_{track_id}.csv from the "
        "previous step. You will be asked once for:\n"
        "  • [bold]Identity cutoff[/bold] (0.0–1.0, default 0.8 = 80%).\n"
        "  • [bold]Method[/bold] — cluster_break / single_linkage / clique."
    )
    outputs_overview = (
        "[bold]CLUSTER_VIEW_{track_id}.csv[/bold]   — slim per-step view (peptide + cluster_id + cluster_size + method).\n"
        "[bold]CLUSTER_{track_id}.csv[/bold]        — full per-peptide row with all upstream columns + cluster info.\n"
        "[bold]CLUSTER_{track_id}.xlsx[/bold]       — same data, sorted by cluster_id with alternating band colours per cluster (orange = cluster_id, pink = cluster_size).\n"
        "[bold]CLUSTER_AUDIT_{track_id}.json[/bold] — threshold, method, n_clusters, n_singletons."
    )
    tips = [
        "80% identity is the default in IEDB tooling and a sensible biological cutoff for MHC-I.",
        "single_linkage is permissive — pick it only if you want to maximise diversity capture.",
        "clique is strict and gives smaller clusters — useful when downstream analysis tolerates redundancy.",
        "Singletons (size=1) are kept and treated as their own clusters — they always become ★ representatives.",
    ]

    def describe_outputs(self) -> dict:
        clusters_dir = self.track_dir / "clusters"
        return {
            clusters_dir / get_step_filename("CLUSTER_VIEW", self.track_id):
                "Slim per-step view — peptide + cluster_id + cluster_size + cluster_method only.",
            clusters_dir / get_step_filename("CLUSTER", self.track_id):
                "Per-peptide clustering result — adds cluster_id, cluster_size, cluster_method columns.",
            clusters_dir / get_step_filename("CLUSTER", self.track_id, ext="xlsx"):
                "Same as the CSV, sorted by cluster_id with alternating band colours so groupings are visible.",
            clusters_dir / get_step_filename("CLUSTER_AUDIT", self.track_id, ext="json"):
                "Run audit — threshold, method, n_clusters, n_singletons.",
        }

    def run(self, input_data=None):
        input_file_path = _find_input_file(self.track_dir, self.track_id)
        epitopes_dataframe = pd.read_csv(input_file_path)

        if COLUMN_PEPTIDE not in epitopes_dataframe.columns:
            raise ValueError(
                f"Column '{COLUMN_PEPTIDE}' not found in {input_file_path.name}."
            )

        epitopes_dataframe[COLUMN_PEPTIDE] = epitopes_dataframe[COLUMN_PEPTIDE].astype(str)
        valid_peptide_mask = epitopes_dataframe[COLUMN_PEPTIDE].apply(
            lambda peptide_sequence: (
                len(peptide_sequence) > 0
                and set(peptide_sequence.upper()) <= _CANONICAL_AMINO_ACIDS
            )
        )
        number_of_dropped_rows = int((~valid_peptide_mask).sum())
        if number_of_dropped_rows:
            console.print(
                f"[yellow]  Dropped {number_of_dropped_rows} row(s) with "
                f"empty or non-canonical peptide residues.[/yellow]"
            )
        epitopes_dataframe = epitopes_dataframe[valid_peptide_mask].reset_index(drop=True)

        epitope_sequences = epitopes_dataframe[COLUMN_PEPTIDE].tolist()
        if not epitope_sequences:
            raise ValueError("No valid peptide sequences found in input file.")

        identity_threshold, clustering_method = _ask_clustering_params(
            self.project_name, self.project_config
        )
        console.print(
            f"  [dim]Clustering {len(epitope_sequences)} epitopes — "
            f"method=[bold]{clustering_method}[/bold]  "
            f"threshold={identity_threshold}[/dim]"
        )

        similarity_matrix = _build_similarity_matrix(epitope_sequences)

        if clustering_method == "clique":
            cluster_labels = _run_clique_clustering(similarity_matrix, identity_threshold)
        elif clustering_method == "single_linkage":
            cluster_labels = _run_single_linkage(similarity_matrix, identity_threshold)
        else:
            cluster_labels = _run_cluster_break(similarity_matrix, identity_threshold)

        epitopes_dataframe["cluster_id"]     = cluster_labels
        epitopes_dataframe["cluster_method"] = clustering_method
        epitopes_dataframe["cluster_size"]   = (
            epitopes_dataframe
            .groupby("cluster_id")["cluster_id"]
            .transform("count")
        )

        clusters_output_dir = self.track_dir / "clusters"
        clusters_output_dir.mkdir(parents=True, exist_ok=True)
        output_csv_path  = clusters_output_dir / get_step_filename("CLUSTER", self.track_id)
        output_xlsx_path = clusters_output_dir / get_step_filename("CLUSTER", self.track_id, ext="xlsx")
        epitopes_dataframe.to_csv(output_csv_path, index=False)
        _write_cluster_xlsx(epitopes_dataframe, output_xlsx_path)

        view_columns = [COLUMN_PEPTIDE, "cluster_id", "cluster_size", "cluster_method"]
        view_csv_path = clusters_output_dir / get_step_filename("CLUSTER_VIEW", self.track_id)
        epitopes_dataframe[view_columns].to_csv(view_csv_path, index=False)

        number_of_clusters  = int(epitopes_dataframe["cluster_id"].nunique())
        number_of_singletons = int((epitopes_dataframe["cluster_size"] == 1).sum())

        audit_data = {
            "timestamp":          datetime.datetime.now().isoformat(),
            "track_id":           self.track_id,
            "input_file":         str(input_file_path),
            "n_epitopes":         len(epitope_sequences),
            "n_dropped_invalid":  number_of_dropped_rows,
            "identity_threshold": identity_threshold,
            "clustering_method":  clustering_method,
            "n_clusters":         number_of_clusters,
            "n_singletons":       number_of_singletons,
        }
        audit_json_path = clusters_output_dir / get_step_filename(
            "CLUSTER_AUDIT", self.track_id, ext="json"
        )
        audit_json_path.write_text(json.dumps(audit_data, indent=2))

        summary_table = Table(box=box.SIMPLE, show_header=False)
        summary_table.add_column(style="cyan")
        summary_table.add_column(justify="right")
        summary_table.add_row("Input file",          input_file_path.name)
        summary_table.add_row("Epitopes clustered",  str(len(epitope_sequences)))
        summary_table.add_row("Method",              clustering_method)
        summary_table.add_row("Identity threshold",  str(identity_threshold))
        summary_table.add_row("Clusters formed",     str(number_of_clusters))
        summary_table.add_row("Singletons",          str(number_of_singletons))
        console.print(summary_table)
        console.print(
            f"\n[bold green]RESULT: {number_of_clusters} clusters from "
            f"{len(epitope_sequences)} epitopes "
            f"(method={clustering_method}, threshold={identity_threshold}).[/bold green]\n"
        )
        console.print(f"  CSV → {output_csv_path}")

        return {
            "cluster_csv":    str(output_csv_path),
            "n_clusters":     number_of_clusters,
            "n_singletons":   number_of_singletons,
        }
