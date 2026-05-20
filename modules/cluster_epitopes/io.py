"""XLSX writer for cluster_epitopes: the CLUSTER sheet with alternating bands."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.naming import COLUMN_PEPTIDE

# ── XLSX writer with alternating cluster bands ────────────────────────────────

_XLSX_FILL_HEADER  = PatternFill("solid", fgColor="D3D3D3")
_XLSX_FILL_BAND_A  = PatternFill("solid", fgColor="FFFFFF")
_XLSX_FILL_BAND_B  = PatternFill("solid", fgColor="EAF3FF")
_XLSX_FILL_CLUSTER = PatternFill("solid", fgColor="FFE4B5")
_XLSX_FILL_SIZE    = PatternFill("solid", fgColor="FFB6C1")
_XLSX_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)


def _write_cluster_xlsx(epitopes_dataframe: pd.DataFrame, output_path: Path) -> None:
    """Cluster table sorted by cluster_id with alternating bands per cluster.
    Header gray bold; body white/light-blue per cluster_id; cluster_id orange, cluster_size pink."""
    sort_columns = [c for c in ("cluster_id", COLUMN_PEPTIDE) if c in epitopes_dataframe.columns]
    if sort_columns:
        df = epitopes_dataframe.sort_values(sort_columns).reset_index(drop=True)
    else:
        df = epitopes_dataframe.reset_index(drop=True)

    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clusters"
    columns = list(df.columns)

    header_font = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.alignment = center
        cell.fill      = _XLSX_FILL_HEADER
        cell.border    = _XLSX_BORDER

    band_toggle = 0
    previous_cluster_id = None
    for row_index, df_row in df.iterrows():
        cluster_id_value = df_row.get("cluster_id")
        if cluster_id_value != previous_cluster_id:
            band_toggle = 1 - band_toggle
            previous_cluster_id = cluster_id_value
        band_fill = _XLSX_FILL_BAND_A if band_toggle == 0 else _XLSX_FILL_BAND_B

        for col_idx, col_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index + 2, column=col_idx, value=df_row[col_name])
            cell.border = _XLSX_BORDER
            if col_name == "cluster_id":
                cell.fill = _XLSX_FILL_CLUSTER
                cell.font = header_font
            elif col_name == "cluster_size":
                cell.fill = _XLSX_FILL_SIZE
            else:
                cell.fill = band_fill

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        cell_widths = [len(str(v)) for v in df[col_name].tolist()]
        max_width = max([len(str(col_name))] + cell_widths)
        worksheet.column_dimensions[col_letter].width = min(max_width + 2, 60)

    worksheet.freeze_panes = "A2"
    workbook.save(str(output_path))


