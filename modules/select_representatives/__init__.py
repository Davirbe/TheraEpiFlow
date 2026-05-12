"""
select_representatives step.

For each cluster produced by cluster_epitopes, selects the single best epitope
representative using a balanced score of binding quality and promiscuity:

  1. best_combined_percentile  — minimum across ALL values in
                                  netmhcpan_el_percentiles_all +
                                  mhcflurry_presentation_percentiles_all
                                  (lower = stronger binder, regardless of allele)
  2. num_alleles_united        — size of the union of netmhcpan_alleles
                                  and mhcflurry_alleles (higher = more promiscuous)

Min-max normalisation over the whole track:
    norm_best_percentile = (max_pct - pct) / (max_pct - min_pct)   [1.0 if all equal]
    norm_alleles         = (n - min_n) / (max_n - min_n)            [1.0 if all equal]

final_score = (norm_best_percentile + norm_alleles) / 2

BEST_REPRESENTATIVE = "★" for the highest final_score per cluster_id.
Ties broken by row order (first row wins).

Input:
    track_dir/clusters/CLUSTER_{track_id}.csv

Output:
    track_dir/clusters/CLUSTER_REPR_{track_id}.csv
    track_dir/clusters/CLUSTER_REPR_{track_id}.xlsx   (colour-coded)
    track_dir/clusters/CLUSTER_REPR_AUDIT_{track_id}.json
"""

import json
import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from rich import box
from utils.console import console
from rich.table import Table

from modules.base_step import BaseTrackStep
from utils.naming import (
    get_step_filename,
    COLUMN_PEPTIDE,
    COLUMN_NETMHC_ALLELES,
    COLUMN_NETMHC_EL_PERCENTILES_ALL,
    COLUMN_FLURRY_ALLELES,
    COLUMN_FLURRY_PERCENTILES_ALL,
    COLUMN_ALLELES_UNITED,
    COLUMN_NUM_ALLELES_UNITED,
    COLUMN_BEST_REPRESENTATIVE,
)

_FILL_ORANGE = PatternFill("solid", fgColor="FFD966")
_FILL_PINK   = PatternFill("solid", fgColor="FFB6C1")
_FILL_YELLOW = PatternFill("solid", fgColor="FFFF99")
_FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")

_PERCENTILE_COLS = frozenset({
    "netmhcpan_el_percentile",
    "netmhcpan_el_percentiles_all",
    "mhcflurry_presentation_percentile",
    "mhcflurry_presentation_percentiles_all",
    "best_combined_percentile",
    "norm_best_percentile",
})
_HLA_COLS = frozenset({
    "netmhcpan_alleles",
    "netmhcpan_num_alleles",
    "mhcflurry_alleles",
    "mhcflurry_num_alleles",
    COLUMN_ALLELES_UNITED,
    COLUMN_NUM_ALLELES_UNITED,
    "norm_alleles",
})


def _parse_semicolon_floats(value) -> list:
    if pd.isna(value) or str(value).strip() == "":
        return []
    try:
        return [float(x) for x in str(value).split(";") if x.strip()]
    except ValueError:
        return []


def _parse_semicolon_strings(value) -> set:
    if pd.isna(value) or str(value).strip() == "":
        return set()
    return {s.strip() for s in str(value).split(";") if s.strip()}


def _min_max_normalize(series: pd.Series, invert: bool = False) -> pd.Series:
    min_val, max_val = series.min(), series.max()
    if max_val == min_val:
        return pd.Series(1.0, index=series.index)
    normalized = (series - min_val) / (max_val - min_val)
    return (1.0 - normalized) if invert else normalized


def _best_combined_percentile(row) -> float:
    vals = (
        _parse_semicolon_floats(row.get(COLUMN_NETMHC_EL_PERCENTILES_ALL, ""))
        + _parse_semicolon_floats(row.get(COLUMN_FLURRY_PERCENTILES_ALL, ""))
    )
    return float(min(vals)) if vals else np.nan


def _alleles_union(row) -> str:
    united = (
        _parse_semicolon_strings(row.get(COLUMN_NETMHC_ALLELES, ""))
        | _parse_semicolon_strings(row.get(COLUMN_FLURRY_ALLELES, ""))
    )
    return ";".join(sorted(united))


def _compute_scores(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["best_combined_percentile"] = df.apply(_best_combined_percentile, axis=1)
    df[COLUMN_ALLELES_UNITED]      = df.apply(_alleles_union, axis=1)
    df[COLUMN_NUM_ALLELES_UNITED]  = df[COLUMN_ALLELES_UNITED].apply(
        lambda v: len([s for s in v.split(";") if s]) if v else 0
    )

    df["norm_best_percentile"] = _min_max_normalize(df["best_combined_percentile"], invert=True)
    df["norm_alleles"]         = _min_max_normalize(df[COLUMN_NUM_ALLELES_UNITED].astype(float))
    df["final_score"]          = (df["norm_best_percentile"] + df["norm_alleles"]) / 2.0

    best_indices = df.groupby("cluster_id")["final_score"].idxmax()
    df[COLUMN_BEST_REPRESENTATIVE] = ""
    df.loc[best_indices, COLUMN_BEST_REPRESENTATIVE] = "★"

    return df


def _write_xlsx(df: pd.DataFrame, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Representatives"
    columns = list(df.columns)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if col_name in _PERCENTILE_COLS:
            cell.fill = _FILL_ORANGE
        elif col_name in _HLA_COLS:
            cell.fill = _FILL_PINK
        else:
            cell.fill = _FILL_HEADER

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_best = row.get(COLUMN_BEST_REPRESENTATIVE) == "★"
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            if is_best:
                cell.fill = _FILL_YELLOW
                if col_name == COLUMN_BEST_REPRESENTATIVE:
                    cell.font = Font(bold=True)

    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        cell_widths = [len(str(v)) for v in df[col_name].tolist()]
        max_width = max([len(str(col_name))] + cell_widths)
        ws.column_dimensions[col_letter].width = min(max_width + 2, 60)

    wb.save(str(path))


class SelectRepresentativesStep(BaseTrackStep):
    step_name   = "select_representatives"
    description = (
        "Picks one representative ★ epitope per cluster by combining HLA "
        "breadth and percentile rank; the ★ marker is the entry point for "
        "every downstream step (conservation, coverage, murine prediction)."
    )

    def describe_outputs(self) -> dict:
        clusters_dir = self.track_dir / "clusters"
        return {
            clusters_dir / get_step_filename("CLUSTER_REPR", self.track_id):
                "Representatives table — best member per cluster marked with ★ in BEST_REPRESENTATIVE.",
            clusters_dir / get_step_filename("CLUSTER_REPR", self.track_id, ext="xlsx"):
                "Same table with colour coding (orange=percentiles, pink=alleles, yellow=★ rows).",
            clusters_dir / get_step_filename("CLUSTER_REPR_AUDIT", self.track_id, ext="json"):
                "Run audit — counts of clusters/singletons/representatives and the scoring rule used.",
        }

    def run(self, input_data=None):
        clusters_dir = self.track_dir / "clusters"
        input_csv = clusters_dir / get_step_filename("CLUSTER", self.track_id)
        if not input_csv.exists():
            raise FileNotFoundError(
                f"cluster_epitopes output not found: {input_csv}\n"
                "Run 'cluster_epitopes' before 'select_representatives'."
            )

        df = pd.read_csv(input_csv)
        if COLUMN_PEPTIDE not in df.columns:
            raise ValueError(f"Column '{COLUMN_PEPTIDE}' not found in {input_csv.name}.")

        df = _compute_scores(df)

        output_csv  = clusters_dir / get_step_filename("CLUSTER_REPR", self.track_id)
        output_xlsx = clusters_dir / get_step_filename("CLUSTER_REPR", self.track_id, ext="xlsx")
        audit_path  = clusters_dir / get_step_filename("CLUSTER_REPR_AUDIT", self.track_id, ext="json")

        df.to_csv(output_csv, index=False)
        _write_xlsx(df, output_xlsx)

        n_epitopes        = len(df)
        n_clusters        = int(df["cluster_id"].nunique())
        n_representatives = int((df[COLUMN_BEST_REPRESENTATIVE] == "★").sum())

        n_multi    = n_clusters - int(df.groupby("cluster_id").size().eq(1).sum())
        n_singletons = n_clusters - n_multi

        summary = Table(box=box.SIMPLE, show_header=False)
        summary.add_column(style="cyan")
        summary.add_column(justify="right")
        summary.add_row("Input epitopes",             str(n_epitopes))
        summary.add_row("Multi-member clusters",      str(n_multi))
        summary.add_row("Singletons",                 str(n_singletons))
        summary.add_row("Representatives (★)",        str(n_representatives))
        console.print(summary)
        console.print(
            f"\n[bold green]RESULT: {n_representatives} cluster representatives selected "
            f"({n_multi} from multi-member clusters, {n_singletons} singletons).[/bold green]\n"
        )

        audit = {
            "timestamp":              datetime.datetime.now().isoformat(),
            "track_id":               self.track_id,
            "input_file":             str(input_csv),
            "n_epitopes":             n_epitopes,
            "n_clusters":             n_clusters,
            "n_multi_member_clusters": n_multi,
            "n_singletons":           n_singletons,
            "n_representatives":      n_representatives,
            "scoring": {
                "best_combined_percentile": (
                    "min of all values in netmhcpan_el_percentiles_all "
                    "+ mhcflurry_presentation_percentiles_all"
                ),
                "norm_best_percentile": "min-max normalised, inverted (lower percentile → higher score)",
                "alleles_united":       "set-union of netmhcpan_alleles and mhcflurry_alleles",
                "norm_alleles":         "min-max normalised num_alleles_united",
                "final_score":          "(norm_best_percentile + norm_alleles) / 2",
            },
            "output_csv":  str(output_csv),
            "output_xlsx": str(output_xlsx),
        }
        audit_path.write_text(json.dumps(audit, indent=2, ensure_ascii=False))

        return {
            "output_csv":        str(output_csv),
            "output_xlsx":       str(output_xlsx),
            "n_epitopes":        n_epitopes,
            "n_clusters":        n_clusters,
            "n_representatives": n_representatives,
        }
