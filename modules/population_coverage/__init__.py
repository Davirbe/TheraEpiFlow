"""
population_coverage step.

Takes the ★ representatives from select_representatives and computes the
population-coverage percentage for each epitope against one or more human
populations from the IEDB allele-frequency database (vendored as a pickle).

The algorithm matches the IEDB Population Coverage tool's diploid model
on a per-epitope basis:

    For each locus the epitope binds:
        q       = Σ frequency(allele)  over the epitope's HLAs in that locus
        p_locus = 1 - (1 - q)²              (probability of cover, diploid)

    Combine independent loci:
        p_total       = 1 - Π (1 - p_locus)
        coverage_pct  = p_total · 100

This is a QUALITATIVE step — no epitopes are removed. Coverage is recorded
on every ★ representative for downstream filtering in the HTML report.

Inputs (clusters/):
    CLUSTER_REPR_{track_id}.csv          ★ representatives + alleles_united

Outputs (coverage/):
    COVERAGE_{track_id}.csv               long format, one row per (peptide × population)
    COVERAGE_{track_id}.xlsx              same data, coloured
    COVERAGE_DETAIL_{population}_{track_id}.csv    IEDB-style per-population CSV
    COVERAGE_HIT_CHART_{population}_{track_id}.png IEDB-style hit chart PNG
    COVERAGE_MATRIX_{track_id}.png        comparative heatmap (≥2 populations only)
    COVERAGE_AUDIT_{track_id}.json        run metadata
"""

import datetime
import json
import pickle
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rich import box
from rich.panel import Panel
from rich.table import Table

from modules.base_step import BaseTrackStep
from utils.console import console, is_interactive_session
from utils.naming import (
    COLUMN_ALLELES_UNITED,
    COLUMN_BEST_REPRESENTATIVE,
    COLUMN_NUM_ALLELES_UNITED,
    COLUMN_PEPTIDE,
    get_step_filename,
)
from utils.project_manager import save_project_config


# ── Database location ─────────────────────────────────────────────────────────

_DB_PATH    = Path(__file__).parent / "data" / "population_genotype_map.p"
_MHC_CLASS  = "I"  # this pipeline is MHC-I only

# Curated short list of populations for the interactive picker. The user can
# always type a free-form name that is validated against the pickle.
_SUGGESTED_POPULATIONS = [
    "World",
    "Europe",
    "East Asia",
    "South Asia",
    "Africa",
    "North America (Native American)",
    "South America",
    "Brazil",
]

# Coverage thresholds for colour bands (purely presentational).
_COVERAGE_HIGH_BAND     = 50.0
_COVERAGE_MODERATE_BAND = 10.0

# ── Colour palette (matches analyze_conservation) ─────────────────────────────

_FILL_GRAY      = PatternFill("solid", fgColor="D9D9D9")
_FILL_HIGH      = PatternFill("solid", fgColor="92D050")
_FILL_MODERATE  = PatternFill("solid", fgColor="FFFF99")
_FILL_LOW       = PatternFill("solid", fgColor="FF9999")

# IEDB-style hit chart palette (matches the user's prototype).
_HIT_CHART_HEADER_DARK  = "#1a3a5c"
_HIT_CHART_HEADER_MED   = "#2e6096"
_HIT_CHART_HEADER_LIGHT = "#3a7abf"
_HIT_CHART_PLUS_COLOR   = "#cc0000"
_HIT_CHART_MINUS_COLOR  = "#1a3a5c"
_HIT_CHART_ROW_EVEN     = "#f0f5fb"
_HIT_CHART_ROW_ODD      = "#ffffff"
_HIT_CHART_COV_HIGH     = "#d4edda"
_HIT_CHART_COV_MED      = "#fff3cd"
_HIT_CHART_COV_LOW      = "#f8d7da"
_HIT_CHART_HIT_BG       = "#e8f5e9"
_HIT_CHART_BORDER       = "#c0c0c0"
_HIT_CHART_TEXT_HEADER  = "white"


# ── Database loading ──────────────────────────────────────────────────────────

def _load_population_database() -> dict:
    """
    Loads the vendored IEDB allele-frequency pickle. Returns the
    `population_coverage` dict only:

        { mhc_class: { population: { locus: [(allele, freq), ...] } } }

    The pickle also contains country_ethnicity and ethnicity tables, loaded
    and discarded because they are not needed by this pipeline.
    """
    with open(_DB_PATH, "rb") as fh:
        population_db    = pickle.load(fh)
        _country_eth_map = pickle.load(fh)  # noqa: F841 — sequenced in pickle stream
        _ethnicity_map   = pickle.load(fh)  # noqa: F841
    return population_db


def _list_available_populations(population_db: dict) -> list[str]:
    """Sorted list of populations available for class I."""
    return sorted(population_db.get(_MHC_CLASS, {}).keys())


# ── Frequency map + coverage math ─────────────────────────────────────────────

def _get_locus(hla_allele: str) -> str:
    """Returns the locus prefix (e.g. 'HLA-A' from 'HLA-A*01:01')."""
    return hla_allele.split("*", 1)[0]


def _build_population_freq_map(population_db: dict, population: str) -> dict[str, float]:
    """
    Builds a {allele: per-locus-normalized frequency} map for the given
    population. If a locus's raw frequencies sum to > 1 in the source data,
    each frequency is divided by the locus sum (so the locus totals to 1).
    Otherwise raw frequencies are kept (already gametic).
    """
    population_block = population_db.get(_MHC_CLASS, {}).get(population, {})
    freq_map: dict[str, float] = {}
    for _locus, allele_freq_pairs in population_block.items():
        locus_total = sum(freq for _, freq in allele_freq_pairs)
        for allele, freq in allele_freq_pairs:
            freq_map[allele] = (freq / locus_total) if locus_total > 1 else freq
    return freq_map


def _compute_epitope_coverage(
    epitope_alleles: set[str],
    freq_map:        dict[str, float],
) -> tuple[float, set[str]]:
    """
    Returns (coverage_pct, matched_alleles).

    coverage_pct is in [0, 100].
    matched_alleles is the subset of epitope_alleles found in freq_map.
    """
    locus_frequency_sum: dict[str, float] = defaultdict(float)
    matched_alleles: set[str] = set()
    for allele in epitope_alleles:
        if allele in freq_map:
            locus_frequency_sum[_get_locus(allele)] += freq_map[allele]
            matched_alleles.add(allele)

    if not locus_frequency_sum:
        return 0.0, matched_alleles

    not_covered_product = 1.0
    for q in locus_frequency_sum.values():
        q = min(q, 1.0)
        p_locus              = 1.0 - (1.0 - q) ** 2
        not_covered_product *= (1.0 - p_locus)

    coverage_pct = round((1.0 - not_covered_product) * 100.0, 2)
    return coverage_pct, matched_alleles


def _parse_alleles_united(raw_value: object) -> set[str]:
    """Splits the semicolon-joined alleles_united field. Returns an empty set
    when the value is missing or NaN."""
    if pd.isna(raw_value):
        return set()
    return {token.strip() for token in str(raw_value).split(";") if token.strip()}


# ── Filename safety helper ────────────────────────────────────────────────────

def _safe_filename_token(value: str) -> str:
    """Sanitises a population name for use in a filename."""
    cleaned: list[str] = []
    for character in value.strip():
        if character.isalnum() or character in {"_", "-"}:
            cleaned.append(character)
        elif character == " ":
            cleaned.append("_")
        # every other character is dropped
    safe_value = "".join(cleaned)
    return safe_value or "Population"


# ── Interactive prompts (preflight) ───────────────────────────────────────────

def _is_non_interactive() -> bool:
    return not sys.stdin.isatty()


def _prompt_populations(project_name: str, project_config: dict, population_db: dict) -> list[str]:
    """Returns the list of populations to compute. Persists into project_config."""
    saved_populations = project_config.get("coverage_populations")

    if _is_non_interactive():
        return list(saved_populations) if saved_populations else ["World"]

    available_populations = _list_available_populations(population_db)

    if saved_populations:
        console.print(Panel(
            f"[bold]Population coverage — saved selection[/bold]\n\n"
            f"Current populations: [cyan]{', '.join(saved_populations)}[/cyan]\n\n"
            "  [cyan][1][/cyan] Keep current selection\n"
            "  [cyan][2][/cyan] Change selection",
            box=box.ROUNDED, title="Setup: population_coverage", title_align="left",
        ))
        while True:
            try:
                choice = input("> ").strip()
            except EOFError:
                choice = "1"
            if choice in ("1", ""):
                return list(saved_populations)
            if choice == "2":
                break
            console.print("[dim]Type 1 or 2.[/dim]")

    console.print(Panel(
        "[bold]Population coverage — pick one or more populations[/bold]\n\n"
        "[dim]The IEDB allele-frequency database covers 239 populations.\n"
        "Type a comma-separated list of numbers (e.g. '1,3,8') or 'other'\n"
        "to type a free-form name validated against the full list.[/dim]\n\n"
        + "\n".join(
            f"  [cyan][{i + 1}][/cyan] {pop_name}"
            for i, pop_name in enumerate(_SUGGESTED_POPULATIONS)
        )
        + "\n  [cyan][other][/cyan] enter a custom population name",
        box=box.ROUNDED, title="Setup: population_coverage", title_align="left",
    ))

    selected_populations: list[str] = []
    while not selected_populations:
        try:
            raw_input_text = input("> ").strip().lower()
        except EOFError:
            raw_input_text = "1"

        if raw_input_text == "other":
            try:
                custom_name = input("Population name (exact): ").strip()
            except EOFError:
                custom_name = ""
            if custom_name in available_populations:
                selected_populations = [custom_name]
            else:
                console.print(
                    f"[red]'{custom_name}' is not in the database.[/red]\n"
                    f"[dim]Hint: search is case-sensitive. Examples:\n"
                    f"{', '.join(available_populations[:8])}, …[/dim]"
                )
            continue

        try:
            indices = [int(token.strip()) for token in raw_input_text.split(",") if token.strip()]
        except ValueError:
            console.print("[red]Invalid input — type numbers separated by commas, or 'other'.[/red]")
            continue

        candidate_populations: list[str] = []
        for idx in indices:
            if 1 <= idx <= len(_SUGGESTED_POPULATIONS):
                candidate = _SUGGESTED_POPULATIONS[idx - 1]
                if candidate in available_populations:
                    candidate_populations.append(candidate)
                else:
                    console.print(
                        f"[yellow]'{candidate}' not in this version of the IEDB pickle, skipping.[/yellow]"
                    )
            else:
                console.print(f"[yellow]Index {idx} out of range, skipping.[/yellow]")

        if candidate_populations:
            seen: set[str] = set()
            selected_populations = [
                pop for pop in candidate_populations if not (pop in seen or seen.add(pop))
            ]

    project_config["coverage_populations"] = selected_populations
    save_project_config(project_name, project_config)
    console.print(
        f"[dim]Populations saved to project_config: "
        f"{', '.join(selected_populations)}[/dim]"
    )
    return selected_populations


def _prompt_coverage_cutoff(project_name: str, project_config: dict) -> Optional[float]:
    """Returns an optional informational cutoff percentage. Persisted."""
    saved_cutoff = project_config.get("coverage_minimum_pct")

    if _is_non_interactive():
        return float(saved_cutoff) if saved_cutoff is not None else None

    if saved_cutoff is not None:
        console.print(
            f"[dim]Coverage cutoff saved earlier: "
            f"[cyan]{saved_cutoff}%[/cyan] (informational only).[/dim]"
        )
        return float(saved_cutoff)

    console.print(Panel(
        "[bold]Coverage cutoff (optional)[/bold]\n\n"
        "[dim]A reference threshold used only to flag low-coverage epitopes\n"
        "in the final HTML report. It does NOT remove anything here.[/dim]\n\n"
        "  [cyan][1][/cyan] Skip (no cutoff)\n"
        "  [cyan][2][/cyan] Set a value",
        box=box.ROUNDED, title="Setup: population_coverage", title_align="left",
    ))
    while True:
        try:
            choice = input("> ").strip()
        except EOFError:
            choice = "1"
        if choice in ("1", ""):
            project_config["coverage_minimum_pct"] = None
            save_project_config(project_name, project_config)
            return None
        if choice == "2":
            try:
                raw_value = input("Cutoff % (e.g. 10): ").strip().replace(",", ".")
            except EOFError:
                raw_value = ""
            try:
                cutoff_value = float(raw_value)
                if 0.0 <= cutoff_value <= 100.0:
                    project_config["coverage_minimum_pct"] = cutoff_value
                    save_project_config(project_name, project_config)
                    return cutoff_value
            except ValueError:
                pass
            console.print("[red]Invalid value. Use a number between 0 and 100.[/red]")
        else:
            console.print("[dim]Type 1 or 2.[/dim]")


# ── CSV / XLSX writers ────────────────────────────────────────────────────────

_SUMMARY_COLUMNS = [
    "peptide",
    "population",
    "mhc_class",
    "coverage_pct",
    "n_hlas_used",
    "n_hlas_in_db",
    "alleles_united",
]


def _coverage_band(coverage_pct: float, cutoff: Optional[float]) -> str:
    """Returns one of 'high', 'moderate', 'low', 'unknown' for colouring."""
    if coverage_pct <= 0.0:
        return "unknown"
    if coverage_pct >= _COVERAGE_HIGH_BAND:
        return "high"
    effective_min = cutoff if cutoff is not None else _COVERAGE_MODERATE_BAND
    if coverage_pct >= effective_min:
        return "moderate"
    return "low"


def write_summary_csv(rows: list[dict], output_path: Path):
    pd.DataFrame(rows, columns=_SUMMARY_COLUMNS).to_csv(output_path, index=False)


def write_summary_xlsx(rows: list[dict], output_path: Path, cutoff: Optional[float]):
    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Coverage Summary"

    for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
        header_cell           = worksheet.cell(row=1, column=col_index, value=column_name)
        header_cell.font      = Font(bold=True)
        header_cell.fill      = _FILL_GRAY
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    band_to_fill = {
        "high":     _FILL_HIGH,
        "moderate": _FILL_MODERATE,
        "low":      _FILL_LOW,
        "unknown":  _FILL_GRAY,
    }

    for row_index, row_data in enumerate(rows, start=2):
        coverage_pct = float(row_data["coverage_pct"])
        band         = _coverage_band(coverage_pct, cutoff)
        for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
            data_cell = worksheet.cell(row=row_index, column=col_index, value=row_data[column_name])
            if column_name == "coverage_pct":
                data_cell.fill          = band_to_fill[band]
                data_cell.number_format = "0.00"
                data_cell.alignment     = Alignment(horizontal="center", vertical="center")
            elif column_name in {"peptide", "population", "alleles_united", "mhc_class"}:
                data_cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                data_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
        col_letter = worksheet.cell(row=1, column=col_index).column_letter
        if rows:
            max_width = max(
                len(column_name),
                *(len(str(row_data.get(column_name, ""))) for row_data in rows),
            )
        else:
            max_width = len(column_name)
        worksheet.column_dimensions[col_letter].width = min(max_width + 2, 50)

    workbook.save(str(output_path))


def write_detail_csv(
    population:       str,
    epitope_results:  list[dict],
    hlas_with_freq:   list[tuple[str, float]],
    output_path:      Path,
):
    """
    IEDB-style per-population CSV: one row per epitope, columns = HLAs (with
    frequency in header), cells '+'/'-', plus 'Epitope set' totals row.
    """
    with open(output_path, "w", encoding="utf-8") as out_fh:
        hla_header_fields = ",".join(
            f'"{allele} ({freq_pct}%)"' for allele, freq_pct in hlas_with_freq
        )
        out_fh.write(f"epitope,coverage_pct,{hla_header_fields},total_hla_hits\n")

        for result in epitope_results:
            symbol_per_hla = [
                "+" if allele in result["matched_alleles"] else "-"
                for allele, _ in hlas_with_freq
            ]
            out_fh.write(
                f"{result['peptide']},{result['coverage_pct']},"
                f"{','.join(symbol_per_hla)},{len(result['matched_alleles'])}\n"
            )

        totals_per_hla = [
            str(sum(1 for r in epitope_results if allele in r["matched_alleles"]))
            for allele, _ in hlas_with_freq
        ]
        out_fh.write(f"Epitope set,,{','.join(totals_per_hla)},\n")


# ── Hit Chart PNG (per population) ────────────────────────────────────────────

def write_hit_chart_png(
    population:        str,
    track_id:          str,
    epitope_results:   list[dict],
    hlas_with_freq:    list[tuple[str, float]],
    cutoff:            Optional[float],
    output_path:       Path,
):
    """English port of the user's prototype, matching IEDB layout."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.patches as mpatches
    import matplotlib.pyplot as plt

    n_epitopes = len(epitope_results)
    n_hlas     = len(hlas_with_freq)
    if n_epitopes == 0 or n_hlas == 0:
        return

    effective_cutoff = cutoff if cutoff is not None else _COVERAGE_MODERATE_BAND

    col_epitope_width  = 1.6
    col_coverage_width = 0.9
    col_hla_width      = 0.72
    row_height         = 0.38
    header_height      = 0.9

    total_width  = col_epitope_width + col_coverage_width + n_hlas * col_hla_width + 0.5
    total_height = header_height + (n_epitopes + 1) * row_height + 1.0

    fig, ax = plt.subplots(figsize=(total_width, total_height))
    ax.set_xlim(0, total_width)
    ax.set_ylim(0, total_height)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    def _draw_cell(
        x_pos: float, y_pos: float, w: float, h: float,
        text: str, background_color: str, text_color: str = "black",
        font_size: float = 7.5, bold: bool = False,
        h_align: str = "center", v_align: str = "center",
    ):
        ax.add_patch(mpatches.Rectangle(
            (x_pos, y_pos), w, h,
            facecolor=background_color, edgecolor=_HIT_CHART_BORDER,
            linewidth=0.5, zorder=1,
        ))
        ax.text(
            x_pos + w / 2, y_pos + h / 2, text,
            ha=h_align, va=v_align,
            fontsize=font_size, color=text_color,
            fontweight=("bold" if bold else "normal"),
            zorder=2, wrap=False, clip_on=True,
        )

    # Title bar
    title_height = 0.5
    y_title_base = total_height - title_height
    _draw_cell(
        0.1, y_title_base, total_width - 0.2, title_height - 0.05,
        f"{track_id}  —  Epitope Coverage  |  Population: {population}  |  Class: MHC {_MHC_CLASS}",
        _HIT_CHART_HEADER_DARK, _HIT_CHART_TEXT_HEADER, font_size=9, bold=True,
    )

    # HLA group sub-header
    y_hla_caption = y_title_base - 0.32
    x_hla_start   = 0.1 + col_epitope_width + col_coverage_width
    _draw_cell(
        x_hla_start, y_hla_caption,
        n_hlas * col_hla_width, 0.28,
        "HLA allele (genotypic frequency %)",
        _HIT_CHART_HEADER_MED, _HIT_CHART_TEXT_HEADER, font_size=8, bold=True,
    )

    # Column headers
    y_header = y_hla_caption - header_height
    x_cursor = 0.1
    _draw_cell(
        x_cursor, y_header, col_epitope_width, header_height,
        "Epitope", _HIT_CHART_HEADER_DARK, _HIT_CHART_TEXT_HEADER,
        font_size=8.5, bold=True,
    )
    x_cursor += col_epitope_width
    _draw_cell(
        x_cursor, y_header, col_coverage_width, header_height,
        f"Coverage\nClass {_MHC_CLASS}",
        _HIT_CHART_HEADER_DARK, _HIT_CHART_TEXT_HEADER, font_size=8, bold=True,
    )
    x_cursor += col_coverage_width
    for hla_index, (allele, frequency_pct) in enumerate(hlas_with_freq):
        short_label = allele.replace("HLA-", "")
        header_bg   = _HIT_CHART_HEADER_MED if hla_index % 2 == 0 else _HIT_CHART_HEADER_LIGHT
        _draw_cell(
            x_cursor, y_header, col_hla_width, header_height,
            f"{short_label}\n({frequency_pct}%)",
            header_bg, _HIT_CHART_TEXT_HEADER, font_size=6.5, bold=True,
        )
        x_cursor += col_hla_width

    # Epitope rows
    for row_index, result in enumerate(epitope_results):
        y_row    = y_header - (row_index + 1) * row_height
        x_cursor = 0.1
        row_bg   = _HIT_CHART_ROW_EVEN if row_index % 2 == 0 else _HIT_CHART_ROW_ODD

        _draw_cell(
            x_cursor, y_row, col_epitope_width, row_height,
            f"#{row_index + 1}: {result['peptide']}",
            row_bg, "black", font_size=7.5, h_align="center",
        )
        x_cursor += col_epitope_width

        coverage_pct = result["coverage_pct"]
        if coverage_pct >= _COVERAGE_HIGH_BAND:
            coverage_bg = _HIT_CHART_COV_HIGH
        elif coverage_pct >= effective_cutoff:
            coverage_bg = _HIT_CHART_COV_MED
        else:
            coverage_bg = _HIT_CHART_COV_LOW

        _draw_cell(
            x_cursor, y_row, col_coverage_width, row_height,
            f"{coverage_pct}%",
            coverage_bg, "black", font_size=8, bold=True,
        )
        x_cursor += col_coverage_width

        for allele, _ in hlas_with_freq:
            is_restricted = allele in result["matched_alleles"]
            symbol        = "+" if is_restricted else "-"
            text_color    = _HIT_CHART_PLUS_COLOR if is_restricted else _HIT_CHART_MINUS_COLOR
            cell_bg       = _HIT_CHART_HIT_BG if is_restricted else row_bg
            _draw_cell(
                x_cursor, y_row, col_hla_width, row_height,
                symbol, cell_bg, text_color, font_size=9, bold=is_restricted,
            )
            x_cursor += col_hla_width

    # Totals row — combined coverage = 1 - Π(1 - p_i) per epitope.
    y_totals = y_header - (n_epitopes + 1) * row_height
    x_cursor = 0.1
    _draw_cell(
        x_cursor, y_totals, col_epitope_width, row_height,
        "Epitope set", _HIT_CHART_HEADER_DARK, _HIT_CHART_TEXT_HEADER,
        font_size=7.5, bold=True,
    )
    x_cursor += col_epitope_width

    not_covered_product = 1.0
    for result in epitope_results:
        not_covered_product *= (1.0 - result["coverage_pct"] / 100.0)
    epitope_set_coverage = round((1.0 - not_covered_product) * 100.0, 2)

    _draw_cell(
        x_cursor, y_totals, col_coverage_width, row_height,
        f"{epitope_set_coverage}%",
        _HIT_CHART_HEADER_DARK, _HIT_CHART_TEXT_HEADER, font_size=7.5, bold=True,
    )
    x_cursor += col_coverage_width

    for allele, _ in hlas_with_freq:
        hit_count = sum(
            1 for result in epitope_results if allele in result["matched_alleles"]
        )
        _draw_cell(
            x_cursor, y_totals, col_hla_width, row_height,
            str(hit_count),
            _HIT_CHART_HEADER_MED, _HIT_CHART_TEXT_HEADER, font_size=7.5, bold=True,
        )
        x_cursor += col_hla_width

    # Legend
    y_legend_line = y_totals - 0.35
    ax.text(
        0.1, y_legend_line,
        "  +  restricted     -  not restricted     "
        f"■ ≥{_COVERAGE_HIGH_BAND:g}%     ■ ≥{effective_cutoff:g}%     ■ <{effective_cutoff:g}%",
        fontsize=7, color="#444444", va="top",
    )
    legend_entries = [
        (2.2, _HIT_CHART_COV_HIGH),
        (3.4, _HIT_CHART_COV_MED),
        (4.6, _HIT_CHART_COV_LOW),
    ]
    for legend_x, legend_color in legend_entries:
        ax.add_patch(mpatches.Rectangle(
            (legend_x, y_legend_line - 0.18), 0.18, 0.18,
            facecolor=legend_color, edgecolor="#888", linewidth=0.5,
        ))

    plt.tight_layout(pad=0.3)
    fig.savefig(str(output_path), dpi=180, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)


# ── Comparative matrix PNG (≥ 2 populations) ──────────────────────────────────

def write_coverage_matrix_png(
    track_id:           str,
    summary_rows:       list[dict],
    populations_order:  list[str],
    output_path:        Path,
):
    """
    Heatmap: rows = epitopes (sorted by mean coverage desc), columns = populations.
    Only generated when at least two populations are present.
    """
    if len(populations_order) < 2:
        return

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    import seaborn as sns

    summary_df = pd.DataFrame(summary_rows)
    pivot      = summary_df.pivot_table(
        index="peptide", columns="population", values="coverage_pct", aggfunc="first"
    ).reindex(columns=populations_order)

    pivot["__mean__"] = pivot[populations_order].mean(axis=1)
    pivot = pivot.sort_values("__mean__", ascending=False).drop(columns="__mean__")

    n_rows = len(pivot)
    if n_rows == 0:
        return

    fig_height = max(4.0, min(22.0, 0.32 * n_rows + 2.0))
    fig_width  = max(4.5, 1.4 * len(populations_order) + 2.0)
    fig, ax    = plt.subplots(figsize=(fig_width, fig_height))

    sns.heatmap(
        pivot.values,
        ax=ax,
        cmap="RdYlGn",
        vmin=0, vmax=100,
        annot=np.array([[f"{v:.1f}%" for v in row] for row in pivot.values]),
        fmt="",
        annot_kws={"size": 7},
        linewidths=0.4, linecolor="#cccccc",
        cbar_kws={"shrink": 0.7, "label": "Coverage (%)"},
        xticklabels=list(pivot.columns),
        yticklabels=list(pivot.index),
    )
    ax.set_title(f"Population coverage — {track_id}", fontsize=11, fontweight="bold", pad=10)
    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.tick_params(axis="y", labelsize=8)
    ax.tick_params(axis="x", labelsize=9)

    plt.tight_layout()
    fig.savefig(str(output_path), dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)


# ── Rich console summary ──────────────────────────────────────────────────────

def _print_console_summary(
    track_id:    str,
    populations: list[str],
    summary_rows: list[dict],
    n_epitopes:   int,
    cutoff:       Optional[float],
):
    summary_df = pd.DataFrame(summary_rows)

    table = Table(
        title=f"Coverage — {track_id}  ({n_epitopes} ★ epitopes)",
        box=box.SIMPLE,
    )
    table.add_column("Population", style="bold")
    table.add_column("Mean coverage", justify="right")
    table.add_column("Max",           justify="right")
    table.add_column("Min",           justify="right")
    table.add_column("≥ cutoff",      justify="right")
    for population in populations:
        rows_for_pop  = summary_df[summary_df["population"] == population]
        coverage_vals = rows_for_pop["coverage_pct"].astype(float)
        if cutoff is not None:
            above_cutoff_text = f"{int((coverage_vals >= cutoff).sum())}/{len(coverage_vals)}"
        else:
            above_cutoff_text = "—"
        table.add_row(
            population,
            f"{coverage_vals.mean():.2f}%" if len(coverage_vals) else "—",
            f"{coverage_vals.max():.2f}%"  if len(coverage_vals) else "—",
            f"{coverage_vals.min():.2f}%"  if len(coverage_vals) else "—",
            above_cutoff_text,
        )
    console.print(table)


# ── Step ──────────────────────────────────────────────────────────────────────

class PopulationCoverageStep(BaseTrackStep):
    step_name   = "population_coverage"
    description = (
        "For each ★ epitope, computes the fraction of one or more human "
        "populations that carries at least one of the epitope's HLA alleles, "
        "using the vendored IEDB allele-frequency database (diploid coverage "
        "model). Qualitative — never removes epitopes."
    )

    @classmethod
    def preflight(
        cls,
        project_name: str,
        project_config: dict,
        track_ids: list[str],
    ) -> Optional[dict]:
        if not track_ids:
            return None
        # Load the DB once at preflight to validate population names against it.
        population_db = _load_population_database()
        _prompt_populations(project_name, project_config, population_db)
        _prompt_coverage_cutoff(project_name, project_config)
        return None

    @classmethod
    def postflight(
        cls,
        project_name: str,
        project_config: dict,
        track_outcomes: dict,
    ) -> None:
        cutoff = project_config.get("coverage_minimum_pct")
        if cutoff is None or not track_outcomes or not is_interactive_session():
            return

        low_coverage_rows: list[tuple[str, str, float]] = []
        for track_id, outcome in track_outcomes.items():
            if outcome.get("status") != "completed":
                continue
            mean_coverage_per_population = outcome.get("mean_coverage_per_population", {})
            for population, mean_value in mean_coverage_per_population.items():
                if float(mean_value) < float(cutoff):
                    low_coverage_rows.append((track_id, population, float(mean_value)))

        if not low_coverage_rows:
            return

        low_table = Table(box=box.SIMPLE, header_style="bold yellow")
        low_table.add_column("Track", style="cyan")
        low_table.add_column("Population")
        low_table.add_column("Mean coverage", justify="right")
        for track_id, population, mean_value in low_coverage_rows:
            low_table.add_row(track_id, population, f"{mean_value:.2f}%")
        console.print(Panel(
            low_table,
            title=f"[bold]Below {cutoff:g}% cutoff (informational)[/bold]",
            border_style="yellow", box=box.ROUNDED,
        ))

    def describe_outputs(self) -> dict[Path, str]:
        coverage_dir = self.track_dir / "coverage"
        described: dict[Path, str] = {
            coverage_dir / get_step_filename("COVERAGE", self.track_id):
                "Long-format coverage: one row per (peptide × population).",
            coverage_dir / get_step_filename("COVERAGE", self.track_id, ext="xlsx"):
                "Same long-format table, coloured by coverage band.",
            coverage_dir / get_step_filename("COVERAGE_AUDIT", self.track_id, ext="json"):
                "Run metadata, populations selected, allele-match stats.",
        }
        populations = self.project_config.get("coverage_populations", ["World"])
        for population in populations:
            safe_name = _safe_filename_token(population)
            described[coverage_dir / f"COVERAGE_DETAIL_{safe_name}_{self.track_id}.csv"] = (
                f"IEDB-style per-epitope-per-HLA table for {population}."
            )
            described[coverage_dir / f"COVERAGE_HIT_CHART_{safe_name}_{self.track_id}.png"] = (
                f"IEDB-style hit chart PNG for {population}."
            )
        if len(populations) >= 2:
            described[coverage_dir / f"COVERAGE_MATRIX_{self.track_id}.png"] = (
                "Comparative heatmap across selected populations."
            )
        return described

    def run(self, input_data=None):
        coverage_dir = self.track_dir / "coverage"
        coverage_dir.mkdir(parents=True, exist_ok=True)

        # ── Load inputs ──────────────────────────────────────────────────────
        cluster_repr_csv = (
            self.track_dir / "clusters" /
            get_step_filename("CLUSTER_REPR", self.track_id)
        )
        if not cluster_repr_csv.exists():
            raise FileNotFoundError(
                f"select_representatives output not found: {cluster_repr_csv}\n"
                "Run 'select_representatives' before 'population_coverage'."
            )

        repr_df = pd.read_csv(cluster_repr_csv)
        star_df = repr_df[repr_df[COLUMN_BEST_REPRESENTATIVE] == "★"].copy()
        if star_df.empty:
            raise ValueError(
                f"No ★ representatives found in {cluster_repr_csv.name}. "
                "Run 'select_representatives' first."
            )

        # ── Resolve runtime configuration ────────────────────────────────────
        population_db = _load_population_database()
        populations: list[str] = list(
            self.project_config.get("coverage_populations") or ["World"]
        )
        cutoff: Optional[float] = self.project_config.get("coverage_minimum_pct")

        valid_populations = set(_list_available_populations(population_db))
        unknown_populations = [p for p in populations if p not in valid_populations]
        if unknown_populations:
            raise ValueError(
                "Unknown populations in project_config: "
                f"{unknown_populations}. Edit project_config.json or rerun preflight."
            )

        console.print(
            f"[dim]→ Computing coverage for {len(star_df)} ★ epitopes "
            f"× {len(populations)} population(s): "
            f"{', '.join(populations)}[/dim]"
        )

        # ── Compute coverage for every (peptide, population) ─────────────────
        summary_rows:           list[dict]              = []
        per_population_details: dict[str, list[dict]]   = {pop: [] for pop in populations}
        per_population_matched: dict[str, int]          = {pop: 0  for pop in populations}
        per_population_missed:  dict[str, int]          = {pop: 0  for pop in populations}

        for population in populations:
            freq_map = _build_population_freq_map(population_db, population)

            for _, peptide_row in star_df.iterrows():
                peptide          = peptide_row[COLUMN_PEPTIDE]
                epitope_alleles  = _parse_alleles_united(
                    peptide_row.get(COLUMN_ALLELES_UNITED, "")
                )
                coverage_pct, matched_alleles = _compute_epitope_coverage(
                    epitope_alleles, freq_map
                )

                summary_rows.append({
                    "peptide":        peptide,
                    "population":     population,
                    "mhc_class":      _MHC_CLASS,
                    "coverage_pct":   coverage_pct,
                    "n_hlas_used":    int(peptide_row.get(COLUMN_NUM_ALLELES_UNITED, len(epitope_alleles)))
                                       if pd.notna(peptide_row.get(COLUMN_NUM_ALLELES_UNITED, None))
                                       else len(epitope_alleles),
                    "n_hlas_in_db":   len(matched_alleles),
                    "alleles_united": ";".join(sorted(epitope_alleles)),
                })

                per_population_details[population].append({
                    "peptide":          peptide,
                    "coverage_pct":     coverage_pct,
                    "all_alleles":      epitope_alleles,
                    "matched_alleles":  matched_alleles,
                })

                per_population_matched[population] += len(matched_alleles)
                per_population_missed[population]  += len(epitope_alleles - matched_alleles)

        # Sort summary by peptide, then by configured population order.
        population_order_index = {p: i for i, p in enumerate(populations)}
        summary_rows.sort(key=lambda r: (r["peptide"], population_order_index[r["population"]]))

        # ── Write outputs ────────────────────────────────────────────────────
        output_csv  = coverage_dir / get_step_filename("COVERAGE", self.track_id)
        output_xlsx = coverage_dir / get_step_filename("COVERAGE", self.track_id, ext="xlsx")
        audit_path  = coverage_dir / get_step_filename("COVERAGE_AUDIT", self.track_id, ext="json")

        write_summary_csv(summary_rows, output_csv)
        write_summary_xlsx(summary_rows, output_xlsx, cutoff)

        detail_csv_paths:    dict[str, Path] = {}
        hit_chart_png_paths: dict[str, Path] = {}

        for population in populations:
            safe_name     = _safe_filename_token(population)
            detail_csv    = coverage_dir / f"COVERAGE_DETAIL_{safe_name}_{self.track_id}.csv"
            hit_chart_png = coverage_dir / f"COVERAGE_HIT_CHART_{safe_name}_{self.track_id}.png"

            # HLAs appearing in any ★ epitope AND present in the population DB,
            # sorted by per-locus-normalized frequency descending (most frequent
            # alleles appear leftmost in the hit chart — matches the prototype).
            freq_map = _build_population_freq_map(population_db, population)
            seen_alleles: set[str] = set()
            for result in per_population_details[population]:
                seen_alleles.update(result["all_alleles"])
            hlas_with_freq = sorted(
                [(allele, round(freq_map[allele] * 100, 2))
                 for allele in seen_alleles if allele in freq_map],
                key=lambda pair: pair[1], reverse=True,
            )

            write_detail_csv(
                population, per_population_details[population], hlas_with_freq, detail_csv,
            )
            write_hit_chart_png(
                population, self.track_id,
                per_population_details[population], hlas_with_freq, cutoff, hit_chart_png,
            )
            detail_csv_paths[population]    = detail_csv
            hit_chart_png_paths[population] = hit_chart_png

        matrix_png_path: Optional[Path] = None
        if len(populations) >= 2:
            matrix_png_path = coverage_dir / f"COVERAGE_MATRIX_{self.track_id}.png"
            write_coverage_matrix_png(
                self.track_id, summary_rows, populations, matrix_png_path,
            )

        # ── Console summary ──────────────────────────────────────────────────
        _print_console_summary(
            self.track_id, populations, summary_rows, len(star_df), cutoff,
        )

        # ── Audit JSON ───────────────────────────────────────────────────────
        summary_df = pd.DataFrame(summary_rows)
        mean_coverage_per_population: dict[str, float] = {}
        for population in populations:
            rows_for_pop = summary_df[summary_df["population"] == population]
            mean_coverage_per_population[population] = round(
                float(rows_for_pop["coverage_pct"].mean()) if len(rows_for_pop) else 0.0, 4
            )

        audit_payload = {
            "timestamp":                       datetime.datetime.now().isoformat(),
            "track_id":                        self.track_id,
            "mhc_class":                       _MHC_CLASS,
            "populations":                     populations,
            "coverage_minimum_pct":            cutoff,
            "n_star_epitopes":                 int(len(star_df)),
            "matched_alleles_per_population":  per_population_matched,
            "missed_alleles_per_population":   per_population_missed,
            "mean_coverage_per_population":    mean_coverage_per_population,
            "outputs": {
                "summary_csv":   str(output_csv),
                "summary_xlsx":  str(output_xlsx),
                "detail_csvs":   {pop: str(path) for pop, path in detail_csv_paths.items()},
                "hit_charts":    {pop: str(path) for pop, path in hit_chart_png_paths.items()},
                "matrix_png":    str(matrix_png_path) if matrix_png_path else None,
                "audit":         str(audit_path),
            },
        }
        audit_path.write_text(json.dumps(audit_payload, indent=2, ensure_ascii=False))

        return {
            "output_csv":                    str(output_csv),
            "output_xlsx":                   str(output_xlsx),
            "populations":                   populations,
            "n_star_epitopes":               int(len(star_df)),
            "mean_coverage_per_population":  mean_coverage_per_population,
        }
