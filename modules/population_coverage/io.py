"""
Data writers for population_coverage: the long-format summary CSV/XLSX, the
slim VIEW CSV and the IEDB-style per-population detail CSV.
"""

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from utils.csv_write import write_user_facing_csv

from .core import _coverage_band

# ── Colour palette (matches analyze_conservation) ─────────────────────────────

_FILL_GRAY      = PatternFill("solid", fgColor="D9D9D9")
_FILL_HIGH      = PatternFill("solid", fgColor="92D050")
_FILL_MODERATE  = PatternFill("solid", fgColor="FFFF99")
_FILL_LOW       = PatternFill("solid", fgColor="FF9999")

_SUMMARY_COLUMNS = [
    "peptide",
    "population",
    "mhc_class",
    "coverage_pct",
    "n_hlas_used",
    "n_hlas_in_db",
    "alleles_united",
]


def write_summary_csv(rows: list[dict], output_path: Path):
    write_user_facing_csv(pd.DataFrame(rows, columns=_SUMMARY_COLUMNS), output_path)


def write_view_csv(rows: list[dict], output_path: Path):
    """Slim VIEW — only the three step-specific columns; drops metadata
    (mhc_class, n_hlas_used, n_hlas_in_db, alleles_united) that already
    lives elsewhere or in the full CSV."""
    view_df = pd.DataFrame(
        [{"peptide": r["peptide"], "population": r["population"], "coverage_pct": r["coverage_pct"]}
         for r in rows]
    )
    write_user_facing_csv(view_df, output_path)


def write_summary_xlsx(rows: list[dict], output_path: Path, cutoff: Optional[float]):
    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = "Coverage Summary"

    for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
        header_cell           = worksheet.cell(row=1, column=col_index, value=column_name)
        header_cell.font      = Font(bold=True)
        header_cell.fill      = _FILL_GRAY
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    band_to_fill = {
        "high":     _FILL_HIGH,
        "moderate": _FILL_MODERATE,
        "low":      _FILL_LOW,
        "unknown":  _FILL_GRAY,
    }

    for row_index, row_data in enumerate(rows, start=2):
        coverage_pct = float(row_data["coverage_pct"])
        band         = _coverage_band(coverage_pct, cutoff)
        for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
            data_cell = worksheet.cell(row=row_index, column=col_index, value=row_data[column_name])
            if column_name == "coverage_pct":
                data_cell.fill          = band_to_fill[band]
                data_cell.number_format = "0.00"
                data_cell.alignment     = Alignment(horizontal="center", vertical="center")
            elif column_name in {"peptide", "population", "alleles_united", "mhc_class"}:
                data_cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                data_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    for col_index, column_name in enumerate(_SUMMARY_COLUMNS, start=1):
        col_letter = worksheet.cell(row=1, column=col_index).column_letter
        if rows:
            max_width = max(
                len(column_name),
                *(len(str(row_data.get(column_name, ""))) for row_data in rows),
            )
        else:
            max_width = len(column_name)
        worksheet.column_dimensions[col_letter].width = min(max_width + 2, 50)

    workbook.save(str(output_path))


def write_detail_csv(
    population:       str,
    epitope_results:  list[dict],
    hlas_with_freq:   list[tuple[str, float]],
    output_path:      Path,
):
    """
    IEDB-style per-population CSV: one row per epitope, columns = HLAs (with
    frequency in header), cells '+'/'-', plus 'Epitope set' totals row.
    """
    with open(output_path, "w", encoding="utf-8") as out_fh:
        hla_header_fields = ",".join(
            f'"{allele} ({freq_pct}%)"' for allele, freq_pct in hlas_with_freq
        )
        out_fh.write(f"epitope,coverage_pct,{hla_header_fields},total_hla_hits\n")

        for result in epitope_results:
            symbol_per_hla = [
                "+" if allele in result["matched_alleles"] else "-"
                for allele, _ in hlas_with_freq
            ]
            out_fh.write(
                f"{result['peptide']},{result['coverage_pct']},"
                f"{','.join(symbol_per_hla)},{len(result['matched_alleles'])}\n"
            )

        totals_per_hla = [
            str(sum(1 for r in epitope_results if allele in r["matched_alleles"]))
            for allele, _ in hlas_with_freq
        ]
        out_fh.write(f"Epitope set,,{','.join(totals_per_hla)},\n")
