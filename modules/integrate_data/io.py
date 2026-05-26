"""integrate_data — XLSX writers.

Two writers share the canonical TheraEpiFlow palette already used by
select_representatives, analyze_conservation, and population_coverage. Kept
here as a private module-level constant block to mirror the per-module
pattern (no shared utils/xlsx_*.py exists yet).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.csv_write import write_user_facing_csv


# ── Palette (mirrors select_representatives + analyze_conservation) ───────────

_FILL_HEADER     = PatternFill("solid", fgColor="D9D9D9")
_FILL_ORANGE     = PatternFill("solid", fgColor="FFE4B5")   # percentile cols
_FILL_PINK       = PatternFill("solid", fgColor="FFB6C1")   # HLA-count cols

# Conservation label palette (matches analyze_conservation/io.py exactly)
_FILL_CONS_PERFECT  = PatternFill("solid", fgColor="00B050")
_FILL_CONS_HIGH     = PatternFill("solid", fgColor="92D050")
_FILL_CONS_MODERATE = PatternFill("solid", fgColor="FFFF99")
_FILL_CONS_LOW      = PatternFill("solid", fgColor="FF9999")
_FILL_CONS_UNKNOWN  = PatternFill("solid", fgColor="D9D9D9")

# Murine label palette (matches predict_murine bands: optimal / good / borderline / non_binder)
_FILL_MURINE_OPTIMAL    = PatternFill("solid", fgColor="00B050")
_FILL_MURINE_GOOD       = PatternFill("solid", fgColor="92D050")
_FILL_MURINE_BORDERLINE = PatternFill("solid", fgColor="FFFF99")
_FILL_MURINE_NON_BINDER = PatternFill("solid", fgColor="FF9999")

# Coverage band palette (matches population_coverage/io.py thresholds)
_FILL_COV_HIGH = PatternFill("solid", fgColor="92D050")   # ≥ 80%
_FILL_COV_MED  = PatternFill("solid", fgColor="FFFF99")   # 50–80%
_FILL_COV_LOW  = PatternFill("solid", fgColor="FF9999")   # < 50%

_THIN_BORDER = Border(
    left   = Side(style="thin", color="C0C0C0"),
    right  = Side(style="thin", color="C0C0C0"),
    top    = Side(style="thin", color="C0C0C0"),
    bottom = Side(style="thin", color="C0C0C0"),
)

_CONSERVATION_LABEL_FILL = {
    'perfect':              _FILL_CONS_PERFECT,
    'high':                 _FILL_CONS_HIGH,
    'moderate':             _FILL_CONS_MODERATE,
    'low':                  _FILL_CONS_LOW,
    'conservation_unknown': _FILL_CONS_UNKNOWN,
}

_MURINE_LABEL_FILL = {
    'optimal':    _FILL_MURINE_OPTIMAL,
    'good':       _FILL_MURINE_GOOD,
    'borderline': _FILL_MURINE_BORDERLINE,
    'non_binder': _FILL_MURINE_NON_BINDER,
}

# Internal column-name sets used for palette decisions; safe to maintain here
# since they describe display behaviour, not data flow.
_PERCENTILE_COLS_HINT = frozenset({
    'netmhcpan_el_percentile',
    'netmhcpan_el_percentiles_all',
    'mhcflurry_presentation_percentile',
    'mhcflurry_presentation_percentiles_all',
    'best_combined_percentile',
    'norm_best_percentile',
    'murine_best_percentile',
})
_HLA_COUNT_COLS_HINT = frozenset({
    'netmhcpan_num_alleles',
    'mhcflurry_num_alleles',
    'num_alleles_united',
    'num_murine_alleles_bound',
})

# Columns that exist purely so the HTML can render rich tooltips; in XLSX
# they are still emitted (so users can audit them) but rendered as narrow
# columns rather than full-width, to avoid drowning the eye.
_THIN_TOOLTIP_COLS = frozenset({
    'alleles_united',
    'conservation_label',
    'murine_best_percentile',
    'num_murine_alleles_bound',
})


def _coverage_fill_for(value: object) -> PatternFill | None:
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return None
    if numeric_value >= 80:
        return _FILL_COV_HIGH
    if numeric_value >= 50:
        return _FILL_COV_MED
    return _FILL_COV_LOW


def _write_header_row(worksheet, columns: list[str], headers: dict[str, str] | None = None) -> None:
    bold_centered = Font(bold=True)
    center        = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(columns, start=1):
        display_name = (headers or {}).get(col_name, col_name)
        cell           = worksheet.cell(row=1, column=col_idx, value=display_name)
        cell.font      = bold_centered
        cell.alignment = center
        cell.fill      = _FILL_HEADER
        cell.border    = _THIN_BORDER


def _autosize_columns(
    worksheet,
    columns: list[str],
    df: pd.DataFrame,
    thin_cols: set[str],
    headers: dict[str, str] | None = None,
) -> None:
    """Auto-fits column widths; thin-tooltip columns get a fixed narrow width."""
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        if col_name in thin_cols:
            worksheet.column_dimensions[col_letter].width = 12
            continue
        display_name = (headers or {}).get(col_name, col_name)
        sample_values = df[col_name].astype(str).tolist() if col_name in df.columns else []
        max_width = max([len(str(display_name))] + [len(v) for v in sample_values])
        worksheet.column_dimensions[col_letter].width = min(max_width + 2, 50)


# ── MASTER_TABLE_FULL ─────────────────────────────────────────────────────────

def write_full_xlsx(full_df: pd.DataFrame, output_path: Path) -> None:
    """Writes MASTER_TABLE_FULL — every column, gray header, freeze A2, no
    cell coloring (the FULL table is a research-grade dump, not a presentation
    surface)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Full"

    columns = list(full_df.columns)
    _write_header_row(ws, columns)

    for row_idx, (_, row) in enumerate(full_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    _autosize_columns(ws, columns, full_df, thin_cols=set())
    wb.save(str(output_path))


# ── MASTER_TABLE_VIEW ─────────────────────────────────────────────────────────

def write_view_xlsx(
    view_df: pd.DataFrame,
    output_path: Path,
    headers: dict[str, str],
) -> None:
    """Writes MASTER_TABLE_VIEW with the canonical palette:
       - conservation_label colored per band
       - murine_label colored per band
       - coverage_<pop> colored per coverage band
       - percentile columns shaded light orange
       - HLA-count columns shaded light pink
       - alleles_united + tooltip-only columns rendered as thin columns
    Headers come from `headers` (internal → display).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master View"

    columns = list(view_df.columns)
    _write_header_row(ws, columns, headers=headers)

    coverage_columns = {c for c in columns if c.startswith('coverage_')}

    for row_idx, (_, row) in enumerate(view_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value       = row[col_name]
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_name == 'conservation_label':
                cell.fill = _CONSERVATION_LABEL_FILL.get(str(value), _FILL_CONS_UNKNOWN)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == 'murine_label':
                cell.fill = _MURINE_LABEL_FILL.get(str(value), _FILL_HEADER)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in coverage_columns:
                coverage_fill = _coverage_fill_for(value)
                if coverage_fill is not None:
                    cell.fill = coverage_fill
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in _PERCENTILE_COLS_HINT:
                cell.fill = _FILL_ORANGE
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in _HLA_COUNT_COLS_HINT:
                cell.fill = _FILL_PINK
                cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A2"
    _autosize_columns(ws, columns, view_df, thin_cols=set(_THIN_TOOLTIP_COLS), headers=headers)
    wb.save(str(output_path))


# ── CSV writer (VIEW sidecar) ─────────────────────────────────────────────────

def write_view_csv(view_df: pd.DataFrame, output_path: Path, headers: dict[str, str]) -> None:
    """Writes the same VIEW as CSV with the display headers.

    Uses the project-wide `write_user_facing_csv` helper, which applies
    utf-8-sig + QUOTE_NONNUMERIC so the file opens cleanly on Excel /
    LibreOffice / Google Sheets in any locale (the semicolon-separated allele
    list would otherwise get split across columns in pt-BR / de-DE).
    """
    renamed = view_df.rename(columns=headers)
    write_user_facing_csv(renamed, output_path)
